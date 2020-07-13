#!/usr/bin/python
# encoding=utf8

import sys
reload(sys)
sys.setdefaultencoding('utf8')
import gi
gi.require_version('Gtk', '3.0')
from gi.repository import Gtk
#from gi.repository import Gdk
import numpy as np
import math
import sqlite3
import pickle
import openpyxl as xl
from openpyxl.worksheet.copier import WorksheetCopy
from time import gmtime, strftime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from itertools import product
import types
import openpyxl
from openpyxl import worksheet
from openpyxl.utils import range_boundaries


def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    Thank you to Sergey Pikhovkin for the fix
    """

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        """ Set merge on a cell range.  Range is a cell range (e.g. A1:E1)
        This is monkeypatched to remove cell deletion bug
        https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
        """
        if not range_string and not all((start_row, start_column, end_row, end_column)):
            msg = "You have to provide a value either for 'coordinate' or for\
            'start_row', 'start_column', 'end_row' *and* 'end_column'"
            raise ValueError(msg)
        elif not range_string:
            range_string = '%s%s:%s%s' % (get_column_letter(start_column),
                                          start_row,
                                          get_column_letter(end_column),
                                          end_row)
        elif ":" not in range_string:
            if COORD_RE.match(range_string):
                return  # Single cell, do nothing
            raise ValueError("Range must be a cell range (e.g. A1:E1)")
        else:
            range_string = range_string.replace('$', '')

        if range_string not in self._merged_cells:
            self._merged_cells.append(range_string)


        # The following is removed by this monkeypatch:

        # min_col, min_row, max_col, max_row = range_boundaries(range_string)
        # rows = range(min_row, max_row+1)
        # cols = range(min_col, max_col+1)
        # cells = product(rows, cols)

        # all but the top-left cell are removed
        #for c in islice(cells, 1, None):
            #if c in self._cells:
                #del self._cells[c]

    # Apply monkey patch
    worksheet.Worksheet.merge_cells = merge_cells
patch_worksheet()



#from matplotlib.backends import backend_tkagg
#import matplotlib.pyplot as plt

import matplotlib.cm as cm
#Possibly this rendering backend is broken currently
import matplotlib.backends.backend_gtk3agg
from matplotlib.backends.backend_gtk3agg import Figure
from matplotlib.backends.backend_gtk3agg import FigureCanvasGTK3Agg as FigureCanvas
#from matplotlib.backends.backend_gtk3cairo import FigureCanvasGTK3Cairo as FigureCanvas


class Signals:

    def analysis_window_hide(self, widget, event):
        clear_all()
        builder.get_object('analysis_window').hide()
        return True

    def on_main_window_destroy(self, widget):
        Gtk.main_quit()

    def on_quit_clicked(self,widget):
        Gtk.main_quit()

    def combinated_analysis(self, widget):
        global analysis_mode, label
        analysis_mode = 'combinated'
        builder.get_object('label2').set_text('Auswertung: ' + label)
        builder.get_object('analysis_window').show_all()

    def zero_analysis(self, widget):
        global analysis_mode
        analysis_mode = 'zero'
        #builder.get_object('entry_track_shift').set_invisible(True)
        builder.get_object('label2').set_text('Auswertung: ' + label)
        builder.get_object('analysis_window').show_all()

    def plan_analysis(self, widget):
        global analysis_mode
        analysis_mode = 'plan'
        builder.get_object('label2').set_text('Auswertung: ' + label)
        builder.get_object('analysis_window').show_all()

    def hide_custom_window(self, widget, event):
        custom_window.hide()
        return True

    def hide_all_frequencies_window(self, widget, event):
        all_frequencies_window.hide()
        return True
    
    def hide_configure_prognose(self, widget, event):
        configure_prognose.hide()
        return True

    def hide_location_dialog(self, widget, event):
        location_dialog.hide()
        return True

    def hide_plot_universal(self, widget, event):
        builder.get_object('plot_universal').hide()
        return True

    def hide_prognose_data_single(self, widget, event):
        builder.get_object('prognose_data_single').hide()
        return True

    def hide_results_single(self, widget, event):
        builder.get_object('results_single').hide()
        return True

    def on_help_clicked(self, widget):
        helpwindow = Gtk.Window()
        #helpwindow.connect("delete-event", Gtk.main_quit)
        helpwindow.set_default_size(250,150)
        hbox = Gtk.Box(spacing=2)
        helpwindow.add(hbox)#label1 = Gtk.Label(str)
        label1 = Gtk.Label("ViP version 1.04 in developement \n developed by \n Rico Reiner Gottschald \n @DB Systemtechnik \n +49 089 13082460 \n rico.gottschald@deutschebahn.com", xalign=0.5)
        #jtype =
        label1.set_justify(Gtk.Justification.CENTER)
        hbox.pack_start(label1, True, True, 0)
        helpwindow.show_all()

    def on_configure_clicked(self,widget):
        configure_prognose.show_all()
        textview_location_preview.set_sensitive(False)

    def on_save_clicked(self,widget):
        filedialog = FileSaveWindow()
        filedialog.show_all()

    def linear_model_import(self, widget):
        filedialog = load_linear_model()
        filedialog.show_all()

    # currently not used
    def on_saveas_clicked(self,widget):
        saveaswindow = Gtk.Window()
        saveaswindow.set_default_size(300, 150)
        hbox = Gtk.Box(spacing=2)
        saveaswindow.add(hbox)  # label1 = Gtk.Label(str)
        label1 = Gtk.Label("noch nicht implementiert \n\n Speicherung der DAtei unter... .",
                           xalign=0.5)
        label1.set_justify(Gtk.Justification.CENTER)
        hbox.pack_start(label1, True, True, 0)
        saveaswindow.show_all()


    def on_open_clicked(self,widget):
        SavedFileOpen().show_all()
        plot_all_results()
        calc_and_print_overall_results()
        #Signals.on_overall_analysis_clicked(self, widget)
        set_menuitems_active()
        fill_configuration_from_file(builder)


    def on_database_clicked(self,widget):
        database_conversation_window.show_all()

    def on_info_clicked(self, widget):
        infowindow = Gtk.Window()
        #helpwindow.connect("delete-event", Gtk.main_quit)
        infowindow.set_default_size(700,250)
        hbox = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=0.2)
        infowindow.add(hbox)#label1 = Gtk.Label(str)
        infolabel = Gtk.Label("Anhaltswerte für zumutbare Innenraumpegel in Anlehnung an 24. BImSchV \n ", xalign=0.5)
        #jtype =
        infolabel.set_justify(Gtk.Justification.CENTER)

        markup = "<span size='large' font_weight='bold'> Anhaltswerte für zumutbare Innenraumpegel in Anlehnung an 24. BImSchV </span>"
        infolabel.set_markup(markup)
        hbox.pack_start(infolabel, True, True, 0)
        textlabel = Gtk.Label("Kategorie 1 \t \t kombinierter Wohn und Schlafraum \n"
                              "Kategorie 2 \t \t Räume die überwiegend zum schlafen genutzt werden \n"
                              "Kategorie 3 \t \t Wohnräume \n"
                              "Kategorie 4 \t \t Behandlungs und Untersuchungsräume in Artzpraxen, Operationsräume, \n"
                              "\t \t \t \t wissenschaftliche Arbeitsräume, Leseräume in Bibliotheken, Unterichtsräume \n"
                              "Kategorie 5 \t \t Konferenz und Vortragsräume, Büroräume, \n"
                              "\t \t \t \t allgemeine Laborräume \n"
                              "Kategorie 6 \t \t Großraumbüros, Schalterräume, Druckerräume von DV-Anlagen, \n"
                              "\t \t \t \t soweit dort ständige Arbeitsplätze vorhanden sind", xalign=0.5)
        textlabel.set_justify(Gtk.Justification.LEFT)
        hbox.pack_start(textlabel, True, True, 0)
        infowindow.show_all()

    def on_future_clicked(self,widget):
        futurewindow = Gtk.Window()
        futurewindow.set_default_size(300, 150)
        hbox = Gtk.Box(spacing=2)
        futurewindow.add(hbox)  # label1 = Gtk.Label(str)
        label1 = Gtk.Label("in kommenden Versionen geplante Features: \n"
                           "\n - iterative Corridorbestimmung"
                           "\n - SqLite Datenbankanbindung"
                           "\n - Excel Export der Auswertung",
                           xalign=0.5)
        label1.set_justify(Gtk.Justification.CENTER)
        hbox.pack_start(label1, True, True, 0)
        futurewindow.show_all()

    def file_open(self, widget):
        global MP5, MP4, MP3, MP2, MP1, MP5Leq, MP4Leq, MP3Leq, MP2Leq, MP1Leq, y_limits
        filedialog = FileChooserWindow()
        filedialog.show_all()
        MP5, MP4, MP3, MP2, MP1, MP5Leq, MP4Leq, MP3Leq, MP2Leq, MP1Leq, y_limits = read_measurement(filename_infile)
        plot_measurement_from_infile(ax1, MP5, MP4, MP3, MP2, MP1, MP5Leq, MP4Leq, MP3Leq, MP2Leq, MP1Leq, y_limits)
        common_ax2(ax2)
        common_ax3(ax3)

    def on_apply1_clicked(self, widget):
        global distMP5, distMP4, distBld, track_shift, LE, LE_Leq, distLv
        distMP5 = float(builder.get_object('entry_MP5').get_text())
        distMP4 = float(builder.get_object('entry_MP4').get_text())
        distBld = float(builder.get_object('entry_dist_bld').get_text())
        if analysis_mode == 'zero':

            track_shift = 0.0
        else:
            track_shift = float(builder.get_object('entry_track_shift').get_text())
        if  distBld >= distMP4:
            LE = MP4
            LE_Leq = MP4Leq
            distLv = distMP4
        elif distBld >= distMP5 and distBld < distMP4:
            LE = MP5
            LE_Leq = MP5Leq
            distLv = distMP5
        else:
            LE = MP5
            LE_Leq = MP5Leq
            distLv = distMP5
        plot_LE(LE, ax2, ticks)
        sw2.queue_draw()
        interact2_set_active()

    def on_apply2_clicked(self, widget):
        global vel_measure, vel_prognose_0, vel_prognose_plan, vel_corr_0, vel_corr_plan, Lv_corr_0, Lv_corr_plan, Lv_corr_0_Leq, Lv_corr_plan_Leq
        vel_measure, vel_prognose_0, vel_prognose_plan = read_entries_interact2()
        if analysis_mode == 'zero':
            vel_corr_0 = calc_velcorr(vel_prognose_0, vel_measure)
            vel_corr_plan = 0.0
        elif analysis_mode == 'plan':
            # some corrections needed
            vel_corr_0 = calc_velcorr(vel_prognose_plan, vel_measure)
            vel_corr_plan = calc_velcorr(vel_prognose_plan, vel_measure)
        else:
            vel_corr_0 = calc_velcorr(vel_prognose_0, vel_measure)
            vel_corr_plan = calc_velcorr(vel_prognose_plan, vel_measure)
        Lv_corr_0, Lv_corr_plan, Lv_corr_0_Leq, Lv_corr_plan_Leq = calc_Lv_corr()
        plot_LE_vel_corr(LE, Lv_corr_0, Lv_corr_plan, vel_corr_0, vel_corr_plan, ax2, ticks)
        sw2.queue_draw()
        interact3_set_active()

    def on_HoeFi_clicked(self, widget):
        deltaLB,  Lv_outsideBld, Lv_outsideBld_prospective, Lv_outsideBld_Leq, Lv_outsideBld_prospective_Leq = calc_HoeFi_attenuation(Lv_corr_0, Lv_corr_plan, Lv_corr_0_Leq, Lv_corr_plan_Leq, distBld, distLv, track_shift)
        plot_attenuated_spec(Lv_outsideBld,  Lv_outsideBld_prospective, ax3, ticks)
        sw3.queue_draw()
        set_transferfunction_combobox_active()
        set_reduction_interact_active()

    def on_linear_model_clicked(self, widget):
        deltaLB, Lv_outsideBld, Lv_outsideBld_prospective, Lv_outsideBld_Leq, Lv_outsideBld_prospective_Leq = calc_attenuation_from_linear_model(Lv_corr_0, Lv_corr_plan, Lv_corr_0_Leq, Lv_corr_plan_Leq, distBld, distLv, track_shift)
        plot_attenuated_spec(Lv_outsideBld,  Lv_outsideBld_prospective, ax3, ticks)
        sw3.queue_draw()
        set_transferfunction_combobox_active()
        set_reduction_interact_active()

    def on_linear_clicked(self, widget):
        global Lv_outsideBld, Lv_outsideBld_prospective
        deltaLB, Lv_outsideBld, Lv_outsideBld_prospective, Lv_outsideBld_Leq, Lv_outsideBld_prospective_Leq =calc_linear_attenuation(MP4, MP5, MP4Leq, MP5Leq, distMP4, distMP5, distBld, track_shift, vel_corr_0, vel_corr_plan)

        plot_attenuated_spec(Lv_outsideBld,  Lv_outsideBld_prospective, ax3, ticks)
        sw3.queue_draw()
        set_transferfunction_combobox_active()
        set_reduction_interact_active()

    def combobox_ceiling_changed(self, widget):
        global ceiling
        ceiling = combobox_ceiling.get_active_text()
        set_transferfunction_interact_active()
        return ceiling

    def combobox_frequency_changed(self, widget):
        global transferfunction, mode
        mode = 'norm'
        if ceiling == 'Holz':
            filename = './src/transferfunctions_wood'
        elif ceiling == 'Beton':
            filename = './src/transferfunctions_concrete'
        freq = combobox_frequency.get_active_text()
        if combobox_frequency.get_active_text() != None:
            transferfunction = read_transferfunctions_from_table(freq, filename)
            plot_transferfunction(transferfunction, freq, ax4, ticks)
            sw4.queue_draw()
            plot_reductions(custom_function, reduction_type, ax5, ticks)
            sw5.queue_draw()
            interact4_set_active()

    def combobox_landuse_changed(self, widget):
        global Au_day, Au_night, Ao_day, Ao_night, Ar_day, Ar_night
        landuse = combobox_landuse.get_active_text()
        Au_day, Au_night, Ao_day, Ao_night, Ar_day, Ar_night = read_4150_conditions_from_table(landuse)

    def combobox_category_changed(self, widget):
        global air_category
        air_category = combobox_air_category.get_active_text()

    def combobox_datasets_changed(self, widget):
        global dataset_entries
        dataset_entry_count = int(combobox_datasets.get_active_text())
        if dataset_entries != []:
            entrytext = []
            for item in dataset_entries:
                entrytext.append(item.get_text())
            #if dataset_entry_count >= len(dataset_entries):
                #dataset_entries = dataset_entries2configure_window(dataset_entry_count, builder)
            dataset_entries = dataset_entries2configure_window(dataset_entry_count, builder)
            if len(dataset_entries) >= len(entrytext):
                for i in range(0, len(entrytext)):
                    dataset_entries[i].set_text(entrytext[i])
            elif len(dataset_entries) < len(entrytext):
                for i in range(0, len(dataset_entries)):
                    dataset_entries[i].set_text(entrytext[i])

            #else:
                #dataset_entries = dataset_entries2configure_window(dataset_entry_count, builder)
        else:
            dataset_entries = dataset_entries2configure_window(dataset_entry_count, builder)

    #def combobox_dataset4analysis_changed(self, widget):
    #    global dataset
    #    dataset = get_text_from_combobox(combobox_dataset4analysis)
    #    button_write_variables2results.set_sensitive(True)

    def on_measurement_from_file_clicked(self, widget):
        global transferfunction, mode
        mode = 'from_file'
        transferfunction = calc_transferfunction_from_file(MP1Leq, Lv_outsideBld_Leq,  vel_corr_0)
        freq = 'aus Messung'
        plot_transferfunction(transferfunction, freq, ax4, ticks)
        sw4.queue_draw()
        interact4_set_active()
        plot_reductions(custom_function, reduction_type, ax5, ticks)
        sw5.queue_draw()

    def on_custom_tfbld_clicked(self, widget):
        global transferfunction, mode
        mode = 'custom'
        load_custom_tfbld()
        freq = 'custom'
        plot_transferfunction(transferfunction, freq, ax4, ticks)
        sw4.queue_draw()
        interact4_set_active()
        plot_reductions(custom_function, reduction_type, ax5, ticks)
        sw5.queue_draw()

    def on_custom_clicked(self, widget):
        custom_window.show_all()
        global reduction_type
        reduction_type = 'custom'

    def on_padded_sleepers_clicked(self, widget):
        global reduction_type
        reduction_type = 'padded_sleeper'
        custom_function = padded_sleepers()
        plot_reductions(custom_function, reduction_type, ax5, ticks)
        sw5.queue_draw()

    def on_slab_track_clicked(self, widget):
        global reduction_type
        reduction_type = 'slab track'
        custom_function = slab_track()
        plot_reductions(custom_function, reduction_type, ax5, ticks)
        sw5.queue_draw()

    def on_switch_clicked(self, widget):
        switch_dialog.show_all()
        global reduction_type
        reduction_type = 'Weiche'


    def view_all_freq_calc(self, widget):
        all_frequencies_window.show_all()

    def on_all_frequencies_single_clicked(self, widget):
        print_all_freq2textview(analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtm_0'], analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtm_plan'],
                                analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBF_max_0'], analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBF_max_plan'],
                                analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtr_day_0'], analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtr_day_plan'],
                                analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtr_night_0'], analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtr_night_plan'],
                                analysis_results['analysis'][label]['all_frequencies_calculation']['all_Lm_day_0'],analysis_results['analysis'][label]['all_frequencies_calculation']['all_Lm_day_plan'],
                                analysis_results['analysis'][label]['all_frequencies_calculation']['all_Lm_night_0'], analysis_results['analysis'][label]['all_frequencies_calculation']['all_Lm_night_plan'],
                                analysis_results['analysis'][label]['all_frequencies_calculation']['all_check4150_string_0'],analysis_results['analysis'][label]['all_frequencies_calculation']['all_check4150_string_plan'],
                                analysis_results['analysis'][label]['all_frequencies_calculation']['all_air_check_0'], analysis_results['analysis'][label]['all_frequencies_calculation']['all_air_check_plan'],
                                analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_KBF_max'],analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_KBFtm'],
                                analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_KBFtr_day'], analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_KBFtr_night'],
                                analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_Lm_day'], analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_Lm_night'])

        all_frequencies_window.show_all()

    def on_prognose_data_single_clicked(self, widget):
        fill_prognose_data_single(label)
        builder.get_object('prognose_data_single').show_all()

    def on_scalar_clicked(self, widget):
        tfwin1.destroy()
        tfwin2.show_all()

    def on_apply3_clicked(self, widget):
        global trains_day_0, trains_night_0, trains_day_plan, trains_night_plan, trainlength_0, trainlength_plan, v_inside_0, v_inside_plan, Lv_inside_0, Lv_inside_plan, KBFtm_0, KBFtm_plan, KBFtr_day_0,\
            KBFtr_day_plan, KBFtr_night_0, KBFtr_night_plan, KBF_max_0, KBF_max_plan, check4150_string_0, check4150_string_plan, air_check_0, air_check_plan, \
            sig_KBF_max, sig_KBFtm, sig_KBFtr_day, sig_KBFtr_night, sig_Lm_day, sig_Lm_night, Lm_day_0, Lm_night_0, Lm_day_plan, Lm_night_plan


        # call
        trains_day_0, trains_night_0, trains_day_plan, trains_night_plan, trainlength_0, trainlength_plan = read_entries_interact3()
        Lv_inside_0, Lv_inside_plan, Lv_inside_0_Leq, Lv_inside_plan_Leq = calc_level_inside(Lv_outsideBld, Lv_outsideBld_prospective, Lv_outsideBld_Leq, Lv_outsideBld_prospective_Leq, transferfunction, custom_function)
        v_inside_0, v_inside_plan = calc_vel_inside(Lv_inside_0, Lv_inside_plan)
        KBF_0, KBF_plan = calc_KBF(v_inside_0, v_inside_plan, terzfrequencies)
        KBFtm_0, KBFtm_plan = calc_KBFtm(KBF_0, KBF_plan)
        KBF_max_0, KBF_max_plan = calc_KBFmax(KBFtm_0, KBFtm_plan, ceiling)
        KBFtr_day_0, KBFtr_night_0, KBFtr_day_plan, KBFtr_night_plan = calc_KBFtr(KBFtm_0, KBFtm_plan, trains_day_0, trains_night_0, trains_day_plan, trains_night_plan)
        Lm_day_0, Lm_night_0, Lm_day_plan, Lm_night_plan = calc_secondary_airborne_sound(Lv_inside_0_Leq, Lv_inside_plan_Leq, ceiling, trainlength_0, trainlength_plan, vel_prognose_0, vel_prognose_plan, trains_day_0, trains_night_0, trains_day_plan, trains_night_plan)
        check4150_string_0, check4150_string_plan = check_4150_conditions(KBF_max_0, KBF_max_plan, KBFtr_day_0,
                                                                          KBFtr_night_0, KBFtr_day_plan, KBFtr_night_plan, Au_day, Au_night, Ao_day, Ao_night, Ar_day, Ar_night)
        air_check_0, air_check_plan = check_sec_airborne_sound(air_category, Lm_day_0, Lm_night_0, Lm_day_plan, Lm_night_plan)
        sig_KBF_max, sig_KBFtm, sig_KBFtr_day, sig_KBFtr_night, sig_Lm_day, sig_Lm_night = calc_significance(KBF_max_0, KBF_max_plan, KBFtm_0, KBFtm_plan, KBFtr_day_0,
                                                                                   KBFtr_night_0, KBFtr_day_plan, KBFtr_night_plan, Lm_day_0, Lm_night_0, Lm_day_plan, Lm_night_plan)
        warnings2textview(warnings)
        plot_level_inside(Lv_inside_0, Lv_inside_plan, ax6, ticks)
        sw6.queue_draw()
        print2textview(Lm_day_0, Lm_night_0, Lm_day_plan, Lm_night_plan, check4150_string_0, check4150_string_plan,
                       air_check_0, air_check_plan, sig_KBF_max, sig_KBFtm, sig_KBFtr_day, sig_KBFtr_night, sig_Lm_day, sig_Lm_night)


        all_KBFtm_0, all_KBFtm_plan, all_KBF_max_0, all_KBF_max_plan, all_KBFtr_day_0, all_KBFtr_day_plan, all_KBFtr_night_0, all_KBFtr_night_plan, all_Lm_day_0, all_Lm_day_plan, all_Lm_night_0, all_Lm_night_plan, all_check4150_string_0, all_check4150_string_plan, all_air_check_0, all_air_check_plan, all_sig_KBF_max, all_sig_KBFtm, all_sig_KBFtr_day, all_sig_KBFtr_night, all_sig_Lm_day, all_sig_Lm_night = calc_all(Lv_outsideBld, Lv_outsideBld_prospective, Lv_outsideBld_Leq, Lv_outsideBld_prospective_Leq, custom_function, terzfrequencies, trains_day_0, trains_night_0, trains_day_plan, trains_night_plan, ceiling, trainlength_0, trainlength_plan, vel_prognose_0, vel_prognose_plan, Au_day, Au_night, Ao_day,
                                                                          Ao_night, Ar_day, Ar_night, air_category)

        print_all_freq2textview(all_KBFtm_0, all_KBFtm_plan, all_KBF_max_0, all_KBF_max_plan, all_KBFtr_day_0,
                                all_KBFtr_day_plan, all_KBFtr_night_0, all_KBFtr_night_plan, all_Lm_day_0, all_Lm_day_plan,
                                all_Lm_night_0, all_Lm_night_plan, all_check4150_string_0, all_check4150_string_plan,
                                all_air_check_0, all_air_check_plan, all_sig_KBF_max, all_sig_KBFtm, all_sig_KBFtr_day,
                                all_sig_KBFtr_night, all_sig_Lm_day, all_sig_Lm_night)

        #set_add_to_results_active()
        button_write_variables2results.set_sensitive(True)

    def on_custom_apply_clicked(self, widget):
        custom_function = read_custom_function()
        custom_window.hide()
        plot_reductions(custom_function, reduction_type, ax5, ticks)
        sw5.queue_draw()

    def on_switch_apply_clicked(self, widget):
        custom_function=calc_switch_reduction()
        switch_dialog.hide()
        plot_reductions(custom_function, reduction_type, ax5, ticks)
        sw5.queue_draw()

    def apply_configuration(self, Widget):
        elements = get_elements_from_entries(dataset_entries)
        glob_notes = get_notes()
        adapt_analysis_results(elements, glob_notes)
        check_configuration_complete(elements)
        analysis_results['metadata'] = {'tracknumber': tracknumber, 'track_loc': track_loc, 'street': street,
                                        'postal': postal,
                                        'location': location, 'state': state, 'country': country, 'landuse':combobox_landuse.get_active_text(),
                                        'glob_notes':glob_notes, 'Au_day':Au_day, 'Au_night':Au_night, 'Ao_day':Ao_day, 'Ao_night':Ao_night,
                                        'Ar_day':Ar_day, 'Ar_night': Ar_night, 'air_category':air_category, 'OPNV':builder.get_object('switch1').get_state()}
        set_menuitems_active()
        fill_buttons()
        set_address_in_results()
        calc_and_print_overall_results()

    def add_location(self, widget):
        location_dialog.show_all()

    def location_apply(self, Widget):
        global tracknumber, track_loc, street, postal, location, state, country
        tracknumber, track_loc, street, postal, location, state, country = read_location_entries()
        location_preview = location4textview(tracknumber, track_loc, street, postal, location, state, country)
        location2textview(location_preview)
        check_address_complete(tracknumber, track_loc, street, postal, location, state, country)

    def plot_coupling_clicked(self, Widget):
        builder.get_object('plot_title').set_text('Gebäudeankopplung')
        plot_bld_coupling(MP3, MP4, ax8)
        builder.get_object('plot_universal').show_all()

    def plot_tfceiling_clicked(self, Widget):
        builder.get_object('plot_title').set_text('Deckenübertragung')
        plot_bld_coupling(MP1, MP3, ax8)
        builder.get_object('plot_universal').show_all()

    def plot_coupling_single_clicked(self, Widget):
        builder.get_object('plot_title').set_text('Gebäudeankopplung' + ' ' + label)
        plot_bld_coupling(np.array(analysis_results['analysis'][label]['MP3']), np.array(analysis_results['analysis'][label]['MP4']), ax8)
        builder.get_object('plot_universal').show_all()

    def on_plot_measurement_clicked(self, widget):
        builder.get_object('plot_title').set_text('Messungen' + ' ' + label)
        plot_measurement(label, ax8)
        builder.get_object('plot_universal').show_all()

    def plot_tfceiling_single_clicked(self, Widget):
        builder.get_object('plot_title').set_text('Deckenübertragung' + ' ' + label)
        plot_bld_coupling(np.array(analysis_results['analysis'][label]['MP1']), np.array(analysis_results['analysis'][label]['MP3']), ax8)
        builder.get_object('plot_universal').show_all()

    def dataset_apply(self, widget):
        variables2results()
        clear_all()
        calc_and_print_overall_results()
        fill_results_single(label)
        plot_all_results()
        analysis_window.hide()

    '''
    def on_overall_analysis_clicked(self, widget):
        main_window.show_all()
        set_result_buttons_invisible()
        fill_buttons()
        if check_for_analyzed_dataset() == True:
            calc_and_print_overall_results()
            plot_all_results()
        else:
            ax7.clear()
            common_ax7(ax7)
            for i in range(24,38):
                clear_textview(builder.get_object('textview' + str(i)))
            for i in range(39, 42):
                clear_textview(builder.get_object('textview' + str(i)))
            for i in range(43, 45):
                clear_textview(builder.get_object('textview' + str(i)))
        if analysis_results['metadata'].keys() != []:
            set_address_in_results()
    '''

    def on_del_0_clicked(self, widget):
        del analysis_results['analysis'][builder.get_object('train_0').get_label()]
        #main_window.hide()
        Signals.on_overall_analysis_clicked(self, widget)

    def on_del_1_clicked(self, widget):
        del analysis_results['analysis'][builder.get_object('train_1').get_label()]
        main_window.hide()
        Signals.on_overall_analysis_clicked(self, widget)

    def on_del_2_clicked(self, widget):
        del analysis_results['analysis'][builder.get_object('train_2').get_label()]
        main_window.hide()
        Signals.on_overall_analysis_clicked(self, widget)

    def on_del_3_clicked(self, widget):
        del analysis_results['analysis'][builder.get_object('train_3').get_label()]
        main_window.hide()
        Signals.on_overall_analysis_clicked(self, widget)

    def on_del_4_clicked(self, widget):
        del analysis_results['analysis'][builder.get_object('train_4').get_label()]
        main_window.hide()
        Signals.on_overall_analysis_clicked(self, widget)

    def on_del_5_clicked(self, widget):
        del analysis_results['analysis'][builder.get_object('train_5').get_label()]
        main_window.hide()
        Signals.on_overall_analysis_clicked(self, widget)

    def on_del_6_clicked(self, widget):
        del analysis_results['analysis'][builder.get_object('train_6').get_label()]
        main_window.hide()
        Signals.on_overall_analysis_clicked(self, widget)

    def on_del_7_clicked(self, widget):
        del analysis_results['analysis'][builder.get_object('train_7').get_label()]
        main_window.hide()
        Signals.on_overall_analysis_clicked(self, widget)

    def on_del_8_clicked(self, widget):
        del analysis_results['analysis'][builder.get_object('train_8').get_label()]
        main_window.hide()
        Signals.on_overall_analysis_clicked(self, widget)

    def on_del_9_clicked(self, widget):
        del analysis_results['analysis'][builder.get_object('train_9').get_label()]
        main_window.hide()
        Signals.on_overall_analysis_clicked(self, widget)

    def on_del_10_clicked(self, widget):
        del analysis_results['analysis'][builder.get_object('train_10').get_label()]
        main_window.hide()
        Signals.on_overall_analysis_clicked(self, widget)

    def on_del_11_clicked(self, widget):
        del analysis_results['analysis'][builder.get_object('train_11').get_label()]
        main_window.hide()
        Signals.on_overall_analysis_clicked(self, widget)

    def on_del_12_clicked(self, widget):
        del analysis_results['analysis'][builder.get_object('train_12').get_label()]
        main_window.hide()
        Signals.on_overall_analysis_clicked(self, widget)

    def on_del_13_clicked(self, widget):
        del analysis_results['analysis'][builder.get_object('train_13').get_label()]
        main_window.hide()
        Signals.on_overall_analysis_clicked(self, widget)

    def on_del_14_clicked(self, widget):
        del analysis_results['analysis'][builder.get_object('train_14').get_label()]
        main_window.hide()
        Signals.on_overall_analysis_clicked(self, widget)

    def on_train_clicked(self, widget):
        global label
        label = widget.get_label()
        clear_results_single()
        fill_results_single(label)
        if analysis_results['analysis'][label].keys() == []:
            builder.get_object('all_frequencies_single').set_sensitive(False)
            builder.get_object('show_measure').set_sensitive(False)
            builder.get_object('skeleton_data').set_sensitive(False)
            builder.get_object('Gebäudeankopplung_single').set_sensitive(False)
            builder.get_object('Deckenübertragung_single').set_sensitive(False)
        else:
            builder.get_object('all_frequencies_single').set_sensitive(True)
            builder.get_object('show_measure').set_sensitive(True)
            builder.get_object('skeleton_data').set_sensitive(True)
            if np.array_equal(np.array(analysis_results['analysis'][label]['MP4']), np.array(analysis_results['analysis'][label]['MP3'])) == False and np.array_equal(np.array(analysis_results['analysis'][label]['MP3']), np.zeros(20)) == False:
                builder.get_object('Gebäudeankopplung_single').set_sensitive(True)
            if np.array_equal(np.array(analysis_results['analysis'][label]['MP1']), np.array(analysis_results['analysis'][label]['MP3'])) == False and np.array_equal(np.array(analysis_results['analysis'][label]['MP3']), np.zeros(20)) == False and np.array_equal(np.array(analysis_results['analysis'][label]['MP1']),np.zeros(20)) == False:
                builder.get_object('Deckenübertragung_single').set_sensitive(True)

        builder.get_object('results_single').show_all()

    def on_export2xls_clicked(self, widget):
        export2xls()

class DB_Signals:

    def hide_database_conversation_window(self, widget, event):
        database_conversation_window.hide()
        return True

    def combobox_sorting_changed(self, widget):
        global sorting
        sorting = combobox_sorting.get_active_text()
        if sorting == 'Projektnummer':
            com = "SELECT project_number FROM measurements"
            label_1.set_text('Projektnummer')
            fill_comboboxes_from_database(combobox_1, com)
        elif sorting == 'Verkehrstyp':
            com = "SELECT train_category FROM measurements"
            label_1.set_text('Verkehrstyp')
            fill_comboboxes_from_database(combobox_1, com)
        elif sorting == 'Zugtyp':
            com = "SELECT train_type FROM measurements"
            label_1.set_text('Zugtyp')
            fill_comboboxes_from_database(combobox_1, com)
        elif sorting == 'Ort':
            com = "SELECT state FROM measurements"
            label_1.set_text('Bundesland')
            fill_comboboxes_from_database(combobox_1, com)
            #com = "SELECT project_number FROM measurements WHERE Id=:Id", {"Id": uId})
        clear_combobox(combobox_2)
        clear_combobox(combobox_3)
        clear_combobox(combobox_4)
        clear_combobox(combobox_5)

    def combobox_1_changed(self, widget):
        global sorting_argument_1
        entry_combobox_1 = get_text_from_combobox(combobox_1)
        if sorting == 'Projektnummer':
            com = "SELECT train_category FROM measurements WHERE project_number=%s" % where_arg4sql(entry_combobox_1)
            label_2.set_text('Verkehrstyp')
            fill_comboboxes_from_database(combobox_2, com)
            sorting_argument_1 = entry_combobox_1
        if sorting == 'Verkehrstyp':
            com = "SELECT train_type FROM measurements WHERE train_category=%s" % where_arg4sql(entry_combobox_1)
            label_2.set_text('Verkehrstyp')
            fill_comboboxes_from_database(combobox_2, com)
            sorting_argument_1 = entry_combobox_1

        clear_combobox(combobox_3)
        clear_combobox(combobox_4)
        clear_combobox(combobox_5)

    def combobox_2_changed(self, widget):
        global sorting_argument_2
        entry_combobox_2 = get_text_from_combobox(combobox_2)
        if sorting == 'Projektnummer':
            com = "SELECT train_type FROM measurements WHERE project_number=%s and train_category=%s" % (where_arg4sql(sorting_argument_1), where_arg4sql(entry_combobox_2))
            label_3.set_text('Zugtyp')
            fill_comboboxes_from_database(combobox_3, com)
            sorting_argument_2 = entry_combobox_2
        clear_combobox(combobox_4)
        clear_combobox(combobox_5)



    def combobox_3_changed(self, widget):
        global sorting_argument_3
        entry_combobox_3 = get_text_from_combobox(combobox_3)
        if sorting == 'Projektnummer':
            com = "SELECT address FROM measurements WHERE project_number=%s and train_category=%s and train_type=%s" % (where_arg4sql(sorting_argument_1), where_arg4sql(sorting_argument_2), where_arg4sql(entry_combobox_3))
            label_4.set_text('Messort')
            fill_comboboxes_from_database(combobox_4, com)
            sorting_argument_3 = entry_combobox_3
        clear_combobox(combobox_5)

    def combobox_4_changed(self, widget):
        entry_combobox_4 = get_text_from_combobox(combobox_4)
        if sorting == 'Projektnummer':
            com = "SELECT track_number, train_type, track, measurement_velocity, ordinal_number FROM measurements WHERE project_number=%s and train_category=%s and train_type=%s and address=%s" % (
            where_arg4sql(sorting_argument_1), where_arg4sql(sorting_argument_2), where_arg4sql(sorting_argument_3), where_arg4sql(entry_combobox_4))
            label_5.set_text('Messung')
            fill_combobox_train_at_location(combobox_5, com)

    def combobox_5_changed(self, widget):
        global MP5, MP4, MP3, MP2, MP1, MP5Leq, MP4Leq, MP3Leq, MP2Leq, MP1Leq
        MP5, MP4, MP3, MP2, MP1, MP5Leq, MP4Leq, MP3Leq, MP2Leq, MP1Leq = get_measurements_from_DB(db_ordinal[combobox_5.get_active()])
        plot_measurement_from_database(ax_DB, sw_DB, ticks, MP5, MP4, MP3, MP2, MP1, MP5Leq, MP4Leq, MP3Leq, MP2Leq, MP1Leq, labels)

    def apply_db_record(self, widget):
        global y_limits, MP5, MP4, MP3, MP2, MP1, MP5Leq, MP4Leq, MP3Leq, MP2Leq, MP1Leq, datalength
        y_limits = ylim_from_db_record(MP5, MP4, MP3, MP2, MP1, MP5Leq, MP4Leq, MP3Leq, MP2Leq, MP1Leq)
        if database_obj.get_object('check_emission').get_active() == True and database_obj.get_object('check_immission').get_active() == False:
            MP5, MP4, MP3, MP2, MP1, MP5Leq, MP4Leq, MP3Leq, MP2Leq, MP1Leq = get_measurements_from_DB(db_ordinal[combobox_5.get_active()])
            MP3  = np.zeros(20)
            MP2 = np.zeros(20)
            MP1 = np.zeros(20)
            MP3Leq = np.zeros(20)
            MP2Leq = np.zeros(20)
            MP1Leq = np.zeros(20)
            datalength = 2
        else:
            MP5, MP4, MP3, MP2, MP1, MP5Leq, MP4Leq, MP3Leq, MP2Leq, MP1Leq = get_measurements_from_DB(db_ordinal[combobox_5.get_active()])
            datalength = 5
        plot_measurement_from_infile(ax1, MP5, MP4, MP3, MP2, MP1, MP5Leq, MP4Leq, MP3Leq, MP2Leq, MP1Leq, y_limits)
        sw1.queue_draw()
        interact1_set_active()
        database_conversation_window.hide()


class FileChooserWindow(Gtk.Window):

    def __init__(self):
        Gtk.Window.__init__(self)
        self.dialog_window()

    def dialog_window(self):
        global filename_infile
        dialog = Gtk.FileChooserDialog("Please choose a file", self,
            Gtk.FileChooserAction.OPEN,
            (Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL,
             Gtk.STOCK_OPEN, Gtk.ResponseType.OK))

        self.add_filters(dialog)

        response = dialog.run()
        if response == Gtk.ResponseType.OK:
             filename_infile = dialog.get_filename()
        elif response == Gtk.ResponseType.CANCEL:
            print("Cancel clicked")

        dialog.destroy()
        Gtk.Window.destroy(self)
        return filename_infile

    def add_filters(self, dialog):
        filter_mes = Gtk.FileFilter()
        filter_mes.set_name("measurements")
        filter_mes.add_pattern("*.mes")
        dialog.add_filter(filter_mes)

        filter_text = Gtk.FileFilter()
        filter_text.set_name("Text files")
        filter_text.add_mime_type("text/plain")
        dialog.add_filter(filter_text)

        filter_py = Gtk.FileFilter()
        filter_py.set_name("Python files")
        filter_py.add_mime_type("text/x-python")
        dialog.add_filter(filter_py)

        filter_any = Gtk.FileFilter()
        filter_any.set_name("Any files")
        filter_any.add_pattern("*")
        dialog.add_filter(filter_any)

class FileSaveWindow(Gtk.Window):

    def __init__(self):
        Gtk.Window.__init__(self)
        self.dialog_window()

    def dialog_window(self):

        dialog = Gtk.FileChooserDialog("Save file", self,
            Gtk.FileChooserAction.SAVE,
            (Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL,
             Gtk.STOCK_SAVE, Gtk.ResponseType.ACCEPT))

        self.add_filters(dialog)
        dialog.set_default_size(800, 400)

        Gtk.FileChooser.set_do_overwrite_confirmation(dialog, True)
        self.user_edited_new_document = True
        if (self.user_edited_new_document):
            Gtk.FileChooser.set_current_name(dialog, "Untitled document.vip")
        else:
            Gtk.FileChooser.set_filename(dialog, self.old_filename)

        response = dialog.run()

        if response == Gtk.ResponseType.ACCEPT:
            filename = Gtk.FileChooser.get_filename(dialog)
            self.save_to_file(filename)

        dialog.destroy()
        Gtk.Window.destroy(self)

    def save_to_file(self, filename):
        # function to write the file to computer
        output = open(filename, 'wb')
        pickle.dump(analysis_results, output)
        output.close()

    def add_filters(self, dialog):
        filter_vip = Gtk.FileFilter()
        filter_vip.set_name("Erschütterungsprognosen")
        filter_vip.add_pattern("*.vip")
        dialog.add_filter(filter_vip)

        filter_mes = Gtk.FileFilter()
        filter_mes.set_name("measurements")
        filter_mes.add_pattern("*.mes")
        dialog.add_filter(filter_mes)

        filter_text = Gtk.FileFilter()
        filter_text.set_name("Text files")
        filter_text.add_mime_type("text/plain")
        dialog.add_filter(filter_text)

        filter_py = Gtk.FileFilter()
        filter_py.set_name("Python files")
        filter_py.add_mime_type("text/x-python")
        dialog.add_filter(filter_py)

        filter_any = Gtk.FileFilter()
        filter_any.set_name("Any files")
        filter_any.add_pattern("*")
        dialog.add_filter(filter_any)

class SavedFileOpen(Gtk.Window):

    def __init__(self):
        Gtk.Window.__init__(self)
        self.dialog_window()

    def dialog_window(self):
        dialog = Gtk.FileChooserDialog("Please choose a file", self,
            Gtk.FileChooserAction.OPEN,
            (Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL,
             Gtk.STOCK_OPEN, Gtk.ResponseType.OK))

        self.add_filters(dialog)

        response = dialog.run()
        if response == Gtk.ResponseType.OK:
            filename = dialog.get_filename()
            self.open_dict(filename)
        elif response == Gtk.ResponseType.CANCEL:
            print("Cancel clicked")
        dialog.destroy()
        Gtk.Window.destroy(self)


    def add_filters(self, dialog):
        filter_vip = Gtk.FileFilter()
        filter_vip.set_name("Erschütterungsprognosen")
        filter_vip.add_pattern("*.vip")
        dialog.add_filter(filter_vip)

        filter_mes = Gtk.FileFilter()
        filter_mes.set_name("measurements")
        filter_mes.add_pattern("*.mes")
        dialog.add_filter(filter_mes)

        filter_text = Gtk.FileFilter()
        filter_text.set_name("Text files")
        filter_text.add_mime_type("text/plain")
        dialog.add_filter(filter_text)

        filter_py = Gtk.FileFilter()
        filter_py.set_name("Python files")
        filter_py.add_mime_type("text/x-python")
        dialog.add_filter(filter_py)

        filter_any = Gtk.FileFilter()
        filter_any.set_name("Any files")
        filter_any.add_pattern("*")
        dialog.add_filter(filter_any)

    def open_dict(self, filename):
        global analysis_results
        file = open(filename, 'rb')
        analysis_results = pickle.load(file)
        file.close()

class load_linear_model(Gtk.Window):

    def __init__(self):
        Gtk.Window.__init__(self)
        self.dialog_window()

    def dialog_window(self):
        dialog = Gtk.FileChooserDialog("Please choose a file", self,
            Gtk.FileChooserAction.OPEN,
            (Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL,
             Gtk.STOCK_OPEN, Gtk.ResponseType.OK))

        self.add_filters(dialog)

        response = dialog.run()
        if response == Gtk.ResponseType.OK:
            filename = dialog.get_filename()
            self.open_dict(filename)
        elif response == Gtk.ResponseType.CANCEL:
            print("Cancel clicked")
        dialog.destroy()
        Gtk.Window.destroy(self)

    def add_filters(self, dialog):
        filter_mod = Gtk.FileFilter()
        filter_mod.set_name("lineares Modell")
        filter_mod.add_pattern("*.mod")
        dialog.add_filter(filter_mod)

        filter_mes = Gtk.FileFilter()
        filter_mes.set_name("measurements")
        filter_mes.add_pattern("*.mes")
        dialog.add_filter(filter_mes)

        filter_text = Gtk.FileFilter()
        filter_text.set_name("Text files")
        filter_text.add_mime_type("text/plain")
        dialog.add_filter(filter_text)

        filter_py = Gtk.FileFilter()
        filter_py.set_name("Python files")
        filter_py.add_mime_type("text/x-python")
        dialog.add_filter(filter_py)

        filter_any = Gtk.FileFilter()
        filter_any.set_name("Any files")
        filter_any.add_pattern("*")
        dialog.add_filter(filter_any)

    def open_dict(self, filename):
        global linear_model, model
        file = open(filename, 'rb')
        linear_model = pickle.load(file)
        model = True
        button_linear_model.set_sensitive(True)
        file.close()

class load_custom_tfbld(Gtk.Window):

    def __init__(self):
        Gtk.Window.__init__(self)
        self.dialog_window()

    def dialog_window(self):
        dialog = Gtk.FileChooserDialog("Please choose a file", self,
            Gtk.FileChooserAction.OPEN,
            (Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL,
             Gtk.STOCK_OPEN, Gtk.ResponseType.OK))

        self.add_filters(dialog)

        response = dialog.run()
        if response == Gtk.ResponseType.OK:
            filename = dialog.get_filename()
            self.open_tf(filename)
        elif response == Gtk.ResponseType.CANCEL:
            print("Cancel clicked")
        dialog.destroy()
        Gtk.Window.destroy(self)

    def add_filters(self, dialog):
        filter_tfbld = Gtk.FileFilter()
        filter_tfbld.set_name("Transferfunction")
        filter_tfbld.add_pattern("*.tf")
        dialog.add_filter(filter_tfbld)

        filter_mod = Gtk.FileFilter()
        filter_mod.set_name("lineares Modell")
        filter_mod.add_pattern("*.mod")
        dialog.add_filter(filter_mod)

        filter_mes = Gtk.FileFilter()
        filter_mes.set_name("measurements")
        filter_mes.add_pattern("*.mes")
        dialog.add_filter(filter_mes)

        filter_text = Gtk.FileFilter()
        filter_text.set_name("Text files")
        filter_text.add_mime_type("text/plain")
        dialog.add_filter(filter_text)

        filter_py = Gtk.FileFilter()
        filter_py.set_name("Python files")
        filter_py.add_mime_type("text/x-python")
        dialog.add_filter(filter_py)

        filter_any = Gtk.FileFilter()
        filter_any.set_name("Any files")
        filter_any.add_pattern("*")
        dialog.add_filter(filter_any)

    def open_tf(self, filename):
        global transferfunction
        file = open(filename, 'rb')
        transferfunction = np.loadtxt(file)
        file.close()


def read_measurement(infile):
    builder.get_object('Gebäudeankopplung').set_sensitive(False)
    builder.get_object('Deckenübertragung').set_sensitive(False)
    data = np.loadtxt(infile)
    global datalength
    # initialize
    datalength = len(data[0, :])
    if datalength == 2:
        #KBFmax
        MP5 = data[:20, 0]
        MP4 = data[:20, 1]
        MP3 = np.zeros(20)
        MP2 = np.zeros(20)
        MP1 = np.zeros(20)
        #Leq
        # Leq
        MP5Leq = data[20:, 0]
        MP4Leq = data[20:, 1]
        MP3Leq = np.zeros(20)
        MP2Leq = np.zeros(20)
        MP1Leq = np.zeros(20)

    elif datalength == 5:
        #KBFmax
        MP5 = data[:20, 0]
        MP4 = data[:20, 1]
        MP3 = data[:20, 2]
        MP2 = data[:20, 3]
        MP1 = data[:20, 4]
        #Leq
        MP5Leq = data[20:, 0]
        MP4Leq = data[20:, 1]
        MP3Leq = data[20:, 2]
        MP2Leq = data[20:, 3]
        MP1Leq = data[20:, 4]

    else:
        MP5 = data[:20, 0]
        MP4 = data[:20, 1]
        MP3 = data[:20, 3]
        MP2 = data[:20, 4]
        MP1 = data[:20, 5]
        # Leq
        MP5Leq = data[20:, 0]
        MP4Leq = data[20:, 1]
        MP3Leq = data[20:, 3]
        MP2Leq = data[20:, 4]
        MP1Leq = data[20:, 5]

    if np.array_equal(MP4, MP3) == False and np.array_equal(MP3,np.zeros(20)) == False:
        builder.get_object('Gebäudeankopplung').set_sensitive(True)
    if np.array_equal(MP1, MP3) == False and np.array_equal(MP3,np.zeros(20)) == False and np.array_equal(MP1,np.zeros(20)) == False:
        builder.get_object('Deckenübertragung').set_sensitive(True)

    y_limits = 0
    for i in range(0, len(data[0, :])):
        if max(data[:, i]) >= y_limits:
            y_limits = max(data[:, i])
    y_limits = y_limits+10
    mod = y_limits % 10
    frac = 10-mod
    y_limits = y_limits+frac

    interact1_set_active()

    if analysis_mode == 'zero':
        builder.get_object('entry_track_shift').set_sensitive(False)

    return MP5, MP4, MP3, MP2, MP1, MP5Leq, MP4Leq, MP3Leq, MP2Leq, MP1Leq, y_limits

def read_transferfunctions_from_table(freq, filename):
    transferfunction = []
    table = np.loadtxt(filename)
    if freq == '8 Hz':
        transferfunction = table[:, 0]
    elif freq == '10 Hz':
        transferfunction = table[:, 1]
    elif freq == '12.5 Hz':
        transferfunction = table[:, 2]
    elif freq == '16 Hz':
        transferfunction = table[:, 3]
    elif freq == '20 Hz':
        transferfunction = table[:, 4]
    elif freq == '25 Hz':
        transferfunction = table[:, 5]
    elif freq == '31.5 Hz':
        transferfunction = table[:, 6]
    elif freq == '40 Hz':
        transferfunction = table[:, 7]
    elif freq == '50 Hz':
        transferfunction = table[:, 8]
    elif freq == '62.5 Hz':
        transferfunction = table[:, 9]
    elif freq == '80 Hz':
        transferfunction = table[:, 10]
    return transferfunction

def read_entries_interact2():
    if analysis_mode == 'zero':
        vel_measure = float(builder.get_object('entry_vel_measure').get_text())
        vel_prognose_0 = float(builder.get_object('entry_vel_zero').get_text())
        vel_prognose_plan = 0.0
    elif analysis_mode == 'plan':
        vel_measure = float(builder.get_object('entry_vel_measure').get_text())
        vel_prognose_0 = 0.0
        vel_prognose_plan = float(builder.get_object('entry_vel_plan').get_text())
    else:
        vel_measure = float(builder.get_object('entry_vel_measure').get_text())
        vel_prognose_0 = float(builder.get_object('entry_vel_zero').get_text())
        vel_prognose_plan = float(builder.get_object('entry_vel_plan').get_text())
    return vel_measure, vel_prognose_0, vel_prognose_plan

def read_entries_interact3():
    trains_day_0 = float(builder.get_object('entry_trains_day_zero').get_text())
    trains_night_0 = float(builder.get_object('entry_trains_night_zero').get_text())
    trains_day_plan = float(builder.get_object('entry_trains_day_plan').get_text())
    trains_night_plan = float(builder.get_object('entry_trains_night_plan').get_text())
    trainlength_0 = float(builder.get_object('entry_trainlength_zero').get_text())
    trainlength_plan = float(builder.get_object('entry_trainlength_plan').get_text())
    return trains_day_0, trains_night_0, trains_day_plan, trains_night_plan, trainlength_0, trainlength_plan

def read_4150_conditions_from_table(landuse):
    A_table = np.loadtxt('./src/Anhaltswerte_DIN4150-2')
    if landuse == 'Industriegebiet':
        Au_day = A_table[0, 0]
        Au_night = A_table[0, 3]
        Ao_day = A_table[0, 1]
        Ao_night = A_table[0, 4]
        Ar_day = A_table[0, 2]
        Ar_night = A_table[0, 5]
    elif landuse == 'Gewerbegebiet':
        Au_day = A_table[1, 0]
        Au_night = A_table[1, 3]
        Ao_day = A_table[1, 1]
        Ao_night = A_table[1, 4]
        Ar_day = A_table[1, 2]
        Ar_night = A_table[1, 5]
    elif landuse == 'Mischgebiet':
        Au_day = A_table[2, 0]
        Au_night = A_table[2, 3]
        Ao_day = A_table[2, 1]
        Ao_night = A_table[2, 4]
        Ar_day = A_table[2, 2]
        Ar_night = A_table[2, 5]
    elif landuse == 'Wohngebiet':
        Au_day = A_table[3, 0]
        Au_night = A_table[3, 3]
        Ao_day = A_table[3, 1]
        Ao_night = A_table[3, 4]
        Ar_day = A_table[3, 2]
        Ar_night = A_table[3, 5]
    elif landuse == 'Sondergebiet':
        Au_day = A_table[4, 0]
        Au_night = A_table[4, 3]
        Ao_day = A_table[4, 1]
        Ao_night = A_table[4, 4]
        Ar_day = A_table[4, 2]
        Ar_night = A_table[4, 5]
    return Au_day, Au_night, Ao_day, Ao_night, Ar_day, Ar_night

def read_category_values_from_table(air_category):
    BImSch_table = np.loadtxt('./src/Innenraumpegel_24.BImSchV')
    if air_category == 'Kat. 1':
        air_day = BImSch_table[0, 1]
        air_night = BImSch_table[0, 2]
    elif air_category == 'Kat. 2':
        air_day = BImSch_table[1, 1]
        air_night = BImSch_table[1, 2]
    elif air_category == 'Kat. 3':
        air_day = BImSch_table[2, 1]
        air_night = BImSch_table[2, 2]
    elif air_category == 'Kat. 4':
        air_day = BImSch_table[3, 1]
        air_night = BImSch_table[3, 2]
    elif air_category == 'Kat. 5':
        air_day = BImSch_table[4, 1]
        air_night = BImSch_table[4, 2]
    elif air_category == 'Kat. 6':
        air_day = BImSch_table[5, 1]
        air_night = BImSch_table[5, 2]
    return air_day, air_night

def read_custom_function():
    global custom_function
    custom_function = []
    for i in range(1, 21):
        entry = custom_reductions.get_object(('entry' + str(i)))
        entry = float(entry.get_text())
        custom_function.append(entry)
    return custom_function

def calc_switch_reduction():
    global custom_function
    custom_function = np.zeros(20)
    dist_switch = float(builder.get_object('dist_switch').get_text())
    switch_reduction = 6 - 5 * np.log10(np.divide(dist_switch, 8))
    custom_function = custom_function + switch_reduction
    return custom_function

def padded_sleepers():
    global custom_function
    custom_function = [-2, -2, -2, -2, -2, -2, -1, -1, -1, -1, -1, -1, -2, -4, -5, -5, -5, -5, -5, -5]
    # MW von Dorothee
    # custom_function = [-3.3, -4.4, -3.5, -2.3, -1.4, -3.9, -3.1, -1.9, -0.8, 0.7, -1.1, -6.8, -10.9, -12.5, -9.8, -10, -11.9, -11, -8.6, -7.4]
    return custom_function

def slab_track():
    global custom_function
    custom_function = [-3.5, -2, -4, -6, -6, -6, -5, -3, 0, 2.5, 2.5, 1, -0.5, -0.5, -4, -7.5, -8.5, -8.5, -8.5, -8.5]
    return custom_function

def plot_measurement_from_infile(ax, MP5, MP4, MP3, MP2, MP1, MP5Leq, MP4Leq, MP3Leq, MP2Leq, MP1Leq, y_limits):
    ax.clear()
    ax.plot(ticks, MP5, label=r'$\mathtt{Emmission MaxHold}$', color='g', linewidth=1.5)
    ax.plot(ticks, MP4, color='g', linewidth=1.5)
    ax.plot(ticks, MP3, label=r'$\mathtt{Immission MaxHold}$', color='r', linewidth=1, linestyle='--')
    ax.plot(ticks, MP2, color='r', linewidth=1, linestyle='--')
    ax.plot(ticks, MP1, color='r', linewidth=1, linestyle='--')
    ax.plot(ticks, MP5Leq, label=r'$\mathtt{Emmission Leq}$', color='m', linewidth=1.5)
    ax.plot(ticks, MP4Leq, color='m', linewidth=1.5)
    ax.plot(ticks, MP3Leq, label=r'$\mathtt{Immission Leq}$', color='c', linewidth=1, linestyle='--')
    ax.plot(ticks, MP2Leq, color='c', linewidth=1, linestyle='--')
    ax.plot(ticks, MP1Leq, color='c', linewidth=1, linestyle='--')
    ax.legend(loc=0, fontsize=8)
    ax.set_xticks(ticks, minor=False)
    ax.set_xticklabels(labels)
    ax.xaxis.grid(True, which='major')
    ax.set_title('Emmissionsspektren der Messung', fontsize=9)
    ax.set_xlabel('Frequenz [Hz]', fontsize=8)
    ax.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize=8)
    ax.tick_params(labelsize=7)
    ax.set_xlim([0, 21])
    ax.set_ylim([0, y_limits])

def plot_LE(LE, ax2, ticks):
    ax2.clear()
    ax2.plot(ticks, LE, label=r'$\mathtt{LE}$')
    ax2.legend(loc=0, fontsize=8)
    ax2.set_xticks(ticks, minor=False)
    ax2.set_xticklabels(labels)
    ax2.xaxis.grid(True, which='major')
    ax2.set_title('Eingangsterzschnellespektrum', fontsize=9)
    ax2.set_xlabel('Frequenz [Hz]', fontsize = 8)
    ax2.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize = 8)
    ax2.tick_params(labelsize =7)
    ax2.set_ylim([0, y_limits])

def plot_LE_vel_corr(LE, Lv_corr_0, Lv_corr_plan, vel_corr_0, vel_corr_plan, ax2, ticks):
    ax2.clear()
    if analysis_mode == 'zero':
        ax2.plot(ticks, LE, label=r'$\mathtt{L_{V_{1}}(f)}$')
        ax2.plot(ticks, Lv_corr_0, label=r'$\mathtt{L_{V_{2}-0}(f)}$', color='green')
    elif analysis_mode == 'plan':
        ax2.plot(ticks, LE, label=r'$\mathtt{L_{V_{1}}(f)}$')
        ax2.plot(ticks, Lv_corr_plan, label=r'$\mathtt{L_{V_{2}-plan}(f)}$', color='red')
    else:
        ax2.plot(ticks, LE, label=r'$\mathtt{L_{V_{1}}(f)}$')
        ax2.plot(ticks, Lv_corr_0, label=r'$\mathtt{L_{V_{2}-0}(f)}$', color='green')
        if vel_corr_0 == vel_corr_plan:
            ax2.plot(ticks, Lv_corr_plan, label=r'$\mathtt{L_{V_{2}-plan}(f)}$', color='red', linestyle='--')
        else:
            ax2.plot(ticks, Lv_corr_plan, label=r'$\mathtt{L_{V_{2}-plan}(f)}$', color='red')
    ax2.legend(loc=0, fontsize=8)
    ax2.set_xticks(ticks, minor=False)
    ax2.set_xticklabels(labels)
    ax2.xaxis.grid(True, which='major')
    ax2.set_title('Eingangsspektrum und Geschwindigkeitskorrektur', fontsize=9)
    ax2.set_xlabel('Frequenz [Hz]', fontsize = 8)
    ax2.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize = 8)
    ax2.tick_params(labelsize =7)
    ax2.set_xlim([0, 21])
    ax2.set_ylim([0, y_limits])

def plot_attenuated_spec(Lv_outsideBld, Lv_outsideBld_prospective, ax3, ticks):
    ax3.clear()
    if analysis_mode == 'zero':
        ax3.plot(ticks, Lv_outsideBld, label=r'$\mathtt{L_{B-0}(f)}$', color='green')
    elif analysis_mode == 'plan':
        ax3.plot(ticks, Lv_outsideBld_prospective, label=r'$\mathtt{L_{B-plan}(f)}$', color='red')
    else:
        ax3.plot(ticks, Lv_outsideBld, label=r'$\mathtt{L_{B-0}(f)}$', color='green')
        ax3.plot(ticks, Lv_outsideBld_prospective, label=r'$\mathtt{L_{B-plan}(f)}$', color='red')
    ax3.legend(loc=0, fontsize=8)
    ax3.set_xticks(ticks, minor=False)
    ax3.set_xticklabels(labels)
    ax3.xaxis.grid(True, which='major')
    ax3.set_title('Pegel vor dem Gebäude', fontsize=9)
    ax3.set_xlabel('Frequenz [Hz]', fontsize = 8)
    ax3.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize = 8)
    ax3.tick_params(labelsize =7)
    ax3.set_xlim([0, 21])
    ax3.set_ylim([0, y_limits])

def plot_transferfunction(transferfunction, freq, ax4, ticks):
    ax4.clear()
    ax4.plot(ticks, transferfunction, label=r'$\mathtt{%s}$' % freq)
    ax4.legend(loc=0, fontsize=8)
    ax4.set_xticks(ticks, minor=False)
    ax4.set_xticklabels(labels)
    ax4.xaxis.grid(True, which='major')
    ax4.set_title('gebäudespezifische Übertragungsfunktion', fontsize=9)
    ax4.set_xlabel('Frequenz [Hz]', fontsize = 8)
    ax4.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize = 8)
    ax4.tick_params(labelsize =7)
    ax4.set_xlim([0, 21])
    ax4.set_ylim([-10, 30])

def plot_reductions(custom_function, reduction_type, ax5, ticks):
    ax5.clear()
    ax5.plot(ticks, custom_function, label=r'$\mathtt{%s}$' % reduction_type)
    ax5.legend(loc=0, fontsize=8)
    ax5.set_xticks(ticks, minor=False)
    ax5.set_xticklabels(labels)
    ax5.xaxis.grid(True, which='major')
    ax5.set_title('Minderungsmaßnahmen', fontsize=9)
    ax5.set_xlabel('Frequenz [Hz]', fontsize=8)
    ax5.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize=8)
    ax5.tick_params(labelsize=7)
    ax5.set_xlim([0, 21])
    ax5.set_ylim([-10, 30])

def plot_level_inside(Lv_inside_0, Lv_inside_plan, ax6, ticks):
    ax6.clear()
    ax6.plot(ticks, Lv_inside_0, label=r'$\mathtt{L_{v-Raum-0}(f)}$', color='green')
    ax6.plot(ticks, Lv_inside_plan, label=r'$\mathtt{L_{v-Raum-plan}(f)}$', color='red')
    ax6.legend(loc=0, fontsize=8)
    ax6.set_xticks(ticks, minor=False)
    ax6.set_xticklabels(labels)
    ax6.xaxis.grid(True, which='major')
    ax6.set_title('Pegel im Gebäude', fontsize=9)
    ax6.set_xlabel('Frequenz [Hz]', fontsize = 8)
    ax6.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize = 8)
    ax6.tick_params(labelsize =7)
    ax6.set_xlim([0, 21])
    ax6.set_ylim([-10, y_limits])

def plot_bld_coupling(MP3, MP4, ax8):
    coupling = MP3/MP4
    ax8.clear()
    ax8.plot(ticks, coupling)#, label=r'$\mathtt{Gebäudeankopplung}$', color='green')
    #ax8.legend(loc=0, fontsize=8)
    ax8.set_xticks(ticks, minor=False)
    ax8.set_xticklabels(labels_extended, rotation=45)
    ax8.xaxis.grid(True, which='major')
    ax8.set_title('Gebäudeankopplung', fontsize=9)
    ax8.set_xlabel('Frequenz [Hz]', fontsize=8)
    ax8.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize=8)
    ax8.tick_params(labelsize=7)
    ax8.set_xlim([0, 21])
    ax8.set_ylim([0,  max(coupling)+5])

def plot_transferfunction_inside(MP1, MP3, ax8):
    transferfunction_inside = MP1/MP3
    ax8.clear()
    ax8.plot(ticks, transferfunction_inside)#, label=r'$\mathtt{Gebäudeankopplung}$', color='green')
    #ax8.legend(loc=0, fontsize=8)
    ax8.set_xticks(ticks, minor=False)
    ax8.set_xticklabels(labels_extended, rotation=45)
    ax8.xaxis.grid(True, which='major')
    ax8.set_title('Gebäudeankopplung', fontsize=9)
    ax8.set_xlabel('Frequenz [Hz]', fontsize=8)
    ax8.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize=8)
    ax8.tick_params(labelsize=7)
    ax8.set_xlim([0, 21])
    ax8.set_ylim([0, max(transferfunction_inside)+5])

def plot_measurement(label, ax8):
    ax8.clear()
    ax8.plot(ticks, analysis_results['analysis'][label]['MP5'], label=r'$\mathtt{Emmission MaxHold}$', color='g', linewidth=1.5)
    ax8.plot(ticks, analysis_results['analysis'][label]['MP4'], color='g', linewidth=1.5)
    ax8.plot(ticks, analysis_results['analysis'][label]['MP3'], label=r'$\mathtt{Immission MaxHold}$', color='r', linewidth=1, linestyle='--')
    ax8.plot(ticks, analysis_results['analysis'][label]['MP2'], color='r', linewidth=1, linestyle='--')
    ax8.plot(ticks, analysis_results['analysis'][label]['MP1'], color='r', linewidth=1, linestyle='--')
    ax8.plot(ticks, analysis_results['analysis'][label]['MP5Leq'], label=r'$\mathtt{Emmission Leq}$', color='m', linewidth=1.5)
    ax8.plot(ticks, analysis_results['analysis'][label]['MP4Leq'], color='m', linewidth=1.5)
    ax8.plot(ticks, analysis_results['analysis'][label]['MP3Leq'], label=r'$\mathtt{Immission Leq}$', color='c', linewidth=1, linestyle='--')
    ax8.plot(ticks, analysis_results['analysis'][label]['MP2Leq'], color='c', linewidth=1, linestyle='--')
    ax8.plot(ticks, analysis_results['analysis'][label]['MP1Leq'], color='c', linewidth=1, linestyle='--')
    ax8.set_xticks(ticks, minor=False)
    ax8.set_xticklabels(labels_extended, rotation=45)
    ax8.xaxis.grid(True, which='major')
    ax8.set_title('Emmissions- /Immissionsmessungen', fontsize=9)
    ax8.set_xlabel('Frequenz [Hz]', fontsize=8)
    ax8.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize=8)
    ax8.tick_params(labelsize=7)
    ax8.legend(loc=0, fontsize=8)
    ax8.set_xlim([0, 21])

def common_ax1(ax1):
    ax1.clear()
    ax1.set_xticks(ticks, minor=False)#plt.xticks(ticks, fontsize=5)
    ax1.set_xticklabels(labels)
    ax1.xaxis.grid(True, which='major')
    ax1.set_title('Emmissionsspektren der Messung', fontsize=9)
    ax1.set_xlabel('Frequenz [Hz]', fontsize = 8)
    ax1.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize = 8)
    ax1.tick_params(labelsize =7)
    ax1.set_xlim([0, 21])

def common_ax2(ax2):
    ax2.clear()
    ax2.set_xticks(ticks, minor=False)
    ax2.set_xticklabels(labels)
    ax2.xaxis.grid(True, which='major')
    ax2.set_title('Eingangsspektrum und Geschwindigkeitskorrektur', fontsize=9)
    ax2.set_xlabel('Frequenz [Hz]', fontsize = 8)
    ax2.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize = 8)
    ax2.tick_params(labelsize =7)
    ax2.set_xlim([0, 21])

def common_ax3(ax3):
    ax3.clear()
    ax3.set_xticks(ticks, minor=False)#plt.xticks(ticks, fontsize=5)
    ax3.set_xticklabels(labels)
    ax3.xaxis.grid(True, which='major')
    ax3.set_title('Pegel vor dem Gebäude', fontsize=9)
    ax3.set_xlabel('Frequenz [Hz]', fontsize = 8)
    ax3.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize = 8)
    ax3.tick_params(labelsize =7)
    ax3.set_xlim([0, 21])

def common_ax4(ax4):
    ax4.clear()
    ax4.set_xticks(ticks, minor=False)
    ax4.set_xticklabels(labels)
    ax4.xaxis.grid(True, which='major')
    ax4.set_title('gebäudespezifische Übertragungsfunktion', fontsize=9)
    ax4.set_xlabel('Frequenz [Hz]', fontsize = 8)
    ax4.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize = 8)
    ax4.tick_params(labelsize =7)
    ax4.set_xlim([0, 21])

def common_ax5(ax5):
    ax5.clear()
    ax5.set_xticks(ticks, minor=False)
    ax5.set_xticklabels(labels)
    ax5.xaxis.grid(True, which='major')
    ax5.set_title('Minderungsmaßnahmen', fontsize=9)
    ax5.set_xlabel('Frequenz [Hz]', fontsize = 8)
    ax5.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize = 8)
    ax5.tick_params(labelsize =7)
    ax5.set_xlim([0, 21])

def common_ax6(ax6):
    ax6.clear()
    ax6.set_xticks(ticks, minor=False)#plt.xticks(ticks, fontsize=5)
    ax6.set_xticklabels(labels)
    ax6.xaxis.grid(True, which='major')
    ax6.set_title('Pegel vor dem Gebäude', fontsize=9)
    ax6.set_xlabel('Frequenz [Hz]', fontsize = 8)
    ax6.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize = 8)
    ax6.tick_params(labelsize =7)
    ax6.set_xlim([0, 21])

def common_ax7(ax7):
    ax7.clear()
    ax7.set_xticks(ticks, minor=False)  # plt.xticks(ticks, fontsize=5)
    ax7.set_xticklabels(labels_extended, rotation=45)
    ax7.xaxis.grid(True, which='major')
    ax7.set_title('Immissionsspektren im Gebäude', fontsize=9)
    ax7.set_xlabel('Frequenz [Hz]', fontsize=8)
    ax7.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize=8)
    ax7.tick_params(labelsize=7)
    ax7.set_xlim([0, 21])

def common_ax8(ax8):
    ax8.clear()
    ax8.set_xticks(ticks, minor=False)  # plt.xticks(ticks, fontsize=5)
    ax8.set_xticklabels(labels_extended, rotation=45)
    ax8.xaxis.grid(True, which='major')
    ax8.set_xlabel('Frequenz [Hz]', fontsize=8)
    ax8.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize=8)
    ax8.tick_params(labelsize=7)
    ax8.set_xlim([0, 21])

def calc_velcorr(vel_prognose, vel_measure):
    global warning_vel
    vel_corr = 20*math.log(vel_prognose/vel_measure, 10)
    if vel_measure*1.3 <= vel_prognose:
        warning_vel = 'Geschwindigkeitskorrektur > 30% der Messgeschwindigkeit nicht zulässig'
    else:
        warning_vel = 0
    return vel_corr

def calc_Lv_corr():
    Lv_corr_0 = LE + vel_corr_0
    Lv_corr_plan = LE + vel_corr_plan
    Lv_corr_0_Leq = LE_Leq + vel_corr_0
    Lv_corr_plan_Leq = LE_Leq + vel_corr_plan
    return Lv_corr_0, Lv_corr_plan, Lv_corr_0_Leq, Lv_corr_plan_Leq

def calc_HoeFi_attenuation(Lv_corr_0, Lv_corr_plan, Lv_corr_0_Leq, Lv_corr_plan_Leq, distBld, distLv, track_shift):
    global deltaLB, Lv_outsideBld, Lv_outsideBld_prospective, Lv_outsideBld_Leq, Lv_outsideBld_prospective_Leq
    deltaLB= np.zeros(20)
    deltaLB_prospective= np.zeros(20)
    distLv_out = distBld - 3
    for i in range(0, 20):
            if i <= 4:
                deltaLB[i] = ( - (15/math.log(35/4)) * math.log((distLv_out)/distLv))
                deltaLB_prospective[i] = ( - (15/math.log(35/4)) * math.log(((distLv_out)+track_shift)/distLv))
            elif i > 4 and  i <= 7:
                deltaLB[i] = ( - (22/math.log(10)) * math.log((distLv_out)/distLv))
                deltaLB_prospective[i] = ( - (22/math.log(10)) * math.log(((distLv_out)+track_shift)/distLv))
            elif i > 7 and  i <= 9:
                deltaLB[i] = ( - (22/math.log(10)) * math.log((distLv_out)/distLv))
                deltaLB_prospective[i] = ( - (22/math.log(10)) * math.log(((distLv_out)+track_shift)/distLv))
            else:
                deltaLB[i] = ( - (15.44/math.log(5/2)) * math.log((distLv_out)/distLv))
                deltaLB_prospective[i] = ( - (15.44/math.log(5/2)) * math.log(((distLv_out)+track_shift)/distLv))

    Lv_outsideBld = Lv_corr_0 + deltaLB
    Lv_outsideBld_Leq = Lv_corr_0_Leq + deltaLB
    Lv_outsideBld_prospective = Lv_corr_plan + deltaLB_prospective
    Lv_outsideBld_prospective_Leq = Lv_corr_plan_Leq + deltaLB_prospective
    return deltaLB, Lv_outsideBld, Lv_outsideBld_prospective, Lv_outsideBld_Leq, Lv_outsideBld_prospective_Leq

def calc_attenuation_from_linear_model(Lv_corr_0, Lv_corr_plan, Lv_corr_0_Leq, Lv_corr_plan_Leq, distBld, distLv, track_shift):
    global deltaLB, Lv_outsideBld, Lv_outsideBld_prospective, Lv_outsideBld_Leq, Lv_outsideBld_prospective_Leq
    deltaLB= np.zeros(20)
    deltaLB_prospective= np.zeros(20)
    distLv_out = distBld - 3.0
    slope = linear_model
    dist_0 = distLv_out - distLv
    dist_plan = (distLv_out + track_shift) - distLv
    for i in range(0, 20):
        # Plan-Fall
        deltaLB[i] = slope[i] * dist_0
        # Plan-Fall
        deltaLB_prospective[i] = slope[i] * dist_plan

    Lv_outsideBld = Lv_corr_0 + deltaLB
    Lv_outsideBld_Leq = Lv_corr_0_Leq + deltaLB
    Lv_outsideBld_prospective = Lv_corr_plan + deltaLB_prospective
    Lv_outsideBld_prospective_Leq = Lv_corr_plan_Leq + deltaLB_prospective

    return deltaLB, Lv_outsideBld, Lv_outsideBld_prospective, Lv_outsideBld_Leq, Lv_outsideBld_prospective_Leq

def calc_linear_attenuation(MP4, MP5, MP4Leq, MP5Leq, distMP4, distMP5, distBld, track_shift, vel_corr_0, vel_corr_plan):
    global deltaLB, Lv_outsideBld, Lv_outsideBld_prospective, Lv_outsideBld_Leq, Lv_outsideBld_prospective_Leq
    distLv_out = distBld-3
    slope = (MP4 - MP5)/(distMP4 - distMP5)
    slope_Leq = (MP4Leq - MP5Leq)/(distMP4 - distMP5)
    # Plan-Fall
    deltaLB = slope*(distLv_out - distMP5)
    Lv_outsideBld = MP5 + deltaLB + vel_corr_0
    deltaLB_Leq = slope_Leq * (distLv_out - distMP5)
    Lv_outsideBld_Leq = MP5Leq + deltaLB_Leq + vel_corr_0
    # Plan-Fall
    deltaLB_prospective = slope * ((distLv_out+track_shift) - distMP5)
    Lv_outsideBld_prospective = MP5 + deltaLB_prospective + vel_corr_plan
    deltaLB_prospective_Leq = slope_Leq * ((distLv_out+track_shift) - distMP5)
    Lv_outsideBld_prospective_Leq = MP5Leq + deltaLB_prospective_Leq + vel_corr_plan
    return deltaLB, Lv_outsideBld, Lv_outsideBld_prospective, Lv_outsideBld_Leq, Lv_outsideBld_prospective_Leq

def calc_transferfunction_from_file(MP1Leq, Lv_outsideBld_Leq, vel_corr_0):
    global transferfunction
    #transferfunction defined 3m point outside the bld to upper ceiling
    v_inside = np.power(10, ((MP1Leq+vel_corr_0)/20))*5*math.pow(10, -5)
    v_outside = np.power(10, (Lv_outsideBld_Leq/20))*5*math.pow(10, -5)

    transferfunction = 20 * np.log10((v_inside/v_outside))

    return transferfunction

def calc_level_inside(Lv_outsideBld, Lv_outsideBld_prospective, Lv_outsideBld_Leq, Lv_outsideBld_prospective_Leq, transferfunction, custom_function):
    Lv_inside_0 = Lv_outsideBld + transferfunction
    Lv_inside_plan = Lv_outsideBld_prospective+transferfunction + custom_function
    Lv_inside_0_Leq = Lv_outsideBld_Leq + transferfunction
    Lv_inside_plan_Leq = Lv_outsideBld_prospective_Leq + transferfunction + custom_function
    return Lv_inside_0, Lv_inside_plan, Lv_inside_0_Leq, Lv_inside_plan_Leq


def calc_vel_inside(Lv_inside_0, Lv_inside_plan):
    v_inside_0 = np.power(10, (Lv_inside_0/20))*(5*math.pow(10, -5))
    v_inside_plan = np.power(10, (Lv_inside_plan/20))*(5*math.pow(10, -5))
    return v_inside_0, v_inside_plan

def calc_KBF(v_inside_0, v_inside_plan, terzfrequencies):
    KBF_0 = v_inside_0/(np.sqrt(1+(np.power(np.divide(5.6, terzfrequencies),2))))
    KBF_plan = v_inside_plan/(np.sqrt(1+(np.power(np.divide(5.6,terzfrequencies),2))))
    return KBF_0, KBF_plan

def calc_KBFtm(KBF_0, KBF_plan):
    sum_KBFtm_0 = 0
    sum_KBFtm_plan = 0
    for i in range(0, 15):
        sum_KBFtm_0 = sum_KBFtm_0 + math.pow(KBF_0[i], 2)
        sum_KBFtm_plan = sum_KBFtm_plan + math.pow(KBF_plan[i], 2)
    KBFtm_0 = math.sqrt(sum_KBFtm_0)
    KBFtm_plan = math.sqrt(sum_KBFtm_plan)
    # 0 setzen gemäß Richtlinie
    #if KBFtm_0 <= 0.1:
    #    KBFtm_0 = 0
    #if KBFtm_plan <= 0.1:
    #    KBFtm_plan = 0
    return KBFtm_0, KBFtm_plan

def calc_KBFmax(KBFtm_0, KBFtm_plan, ceiling):
    KBF_max_0 = []
    KBF_max_plan = []
    if ceiling == 'Holz':
        KBF_max_0 = 1.7*KBFtm_0
        KBF_max_plan = 1.7*KBFtm_plan
    elif ceiling == 'Beton':
        KBF_max_0 = 1.5*KBFtm_0
        KBF_max_plan = 1.5*KBFtm_plan
    return KBF_max_0, KBF_max_plan

def calc_KBFtr(KBFtm_0, KBFtm_plan, trains_day_0, trains_night_0, trains_day_plan, trains_night_plan):
    if KBFtm_0 > KBFtm_plan:
        if KBFtm_0 <= 0.1:
            KBFtr_day_0 = 0
            KBFtr_night_0 = 0
            KBFtr_day_plan = 0
            KBFtr_night_plan = 0
        else:
            KBFtr_day_0 =  math.sqrt(math.pow(KBFtm_0, 2)*(trains_day_0*30/57600))
            KBFtr_night_0 = math.sqrt(math.pow(KBFtm_0, 2)*(trains_night_0*30/28800))
            KBFtr_day_plan = math.sqrt(math.pow(KBFtm_plan, 2)*(trains_day_plan*30/57600))
            KBFtr_night_plan = math.sqrt(math.pow(KBFtm_plan, 2)*(trains_night_plan*30/28800))
    else:
        if KBFtm_plan <= 0.1:
            KBFtr_day_0 = 0
            KBFtr_night_0 = 0
            KBFtr_day_plan = 0
            KBFtr_night_plan = 0
        else:
            KBFtr_day_0 = math.sqrt(math.pow(KBFtm_0, 2) * (trains_day_0 * 30 / 57600))
            KBFtr_night_0 = math.sqrt(math.pow(KBFtm_0, 2) * (trains_night_0 * 30 / 28800))
            KBFtr_day_plan = math.sqrt(math.pow(KBFtm_plan, 2) * (trains_day_plan * 30 / 57600))
            KBFtr_night_plan = math.sqrt(math.pow(KBFtm_plan, 2) * (trains_night_plan * 30 / 28800))
    return KBFtr_day_0, KBFtr_night_0, KBFtr_day_plan, KBFtr_night_plan

def calc_secondary_airborne_sound(Lv_inside_0_Leq, Lv_inside_plan_Leq, ceiling, trainlength_0, trainlength_plan, vel_prognose_0, velprognose_plan, trains_day_0, trains_night_0, trains_day_plan, trains_night_plan):
    # scalar method
    A_rating_scalar_table = np.loadtxt('./src/a_rating_reduced')
    A_rating_scalar = A_rating_scalar_table[:, 1]
    reduced_Lv_inside_0_scalar = Lv_inside_0_Leq[8:15]
    reduced_Lv_inside_plan_scalar = Lv_inside_plan_Leq[8:15]
    reduced_Lv_inside_A_0_scalar = reduced_Lv_inside_0_scalar + A_rating_scalar
    reduced_Lv_inside_A_plan_scalar = reduced_Lv_inside_plan_scalar + A_rating_scalar
    L_sek_sum_0_scalar = 10 * math.log10(np.sum(np.power(10, np.divide(reduced_Lv_inside_A_0_scalar,10))))  # pegel addition over 10 - 100Hz
    L_sek_sum_plan_scalar = 10 * math.log10(np.sum(np.power(10, np.divide(reduced_Lv_inside_A_plan_scalar, 10))))  # pegel addition over 10 - 100Hz

    if ceiling == 'Beton':
        L_sek_0 = 15.75+0.6* L_sek_sum_0_scalar
        L_sek_plan = 15.75 + 0.6 * L_sek_sum_plan_scalar
    elif ceiling == 'Holz':
        L_sek_0 = 19.88 + 0.47 * L_sek_sum_0_scalar
        L_sek_plan = 19.88 + 0.47 * L_sek_sum_plan_scalar

    if analysis_mode == 'zero':
        passingtime_0 = trainlength_0 / (vel_prognose_0 / 3.6)

        if trains_day_0 == 0:
            Lm_day_0 = -100.0
        else:
            Lm_day_0 = L_sek_0 + 10 * math.log((passingtime_0 * trains_day_0 / 57600), 10)

        if trains_night_0 == 0:
            Lm_night_0 = -100.0
        else:
            Lm_night_0 = L_sek_0 + 10 * math.log((passingtime_0 * trains_night_0 / 28800), 10)

        Lm_day_plan = -100.0
        Lm_night_plan = -100.0

    elif analysis_mode == 'plan':

        passingtime_plan = trainlength_plan / (vel_prognose_plan / 3.6)

        if trains_day_plan == 0:
            Lm_day_plan = -100.0
        else:
            Lm_day_plan = L_sek_plan + 10 * math.log((passingtime_plan * trains_day_plan / 57600), 10)

        if trains_night_plan == 0:
            Lm_night_plan = -100.0
        else:
            Lm_night_plan = L_sek_plan + 10 * math.log((passingtime_plan * trains_night_plan / 28800), 10)

        Lm_day_0 = -100.0
        Lm_night_0 = -100.0

    else:
        passingtime_0 = trainlength_0 / (vel_prognose_0 / 3.6)
        passingtime_plan = trainlength_plan / (vel_prognose_plan / 3.6)

        if trains_day_0 == 0:
            Lm_day_0 = -100.0
        else:
            Lm_day_0 = L_sek_0 + 10 * math.log((passingtime_0 * trains_day_0 / 57600), 10)
        if trains_day_plan  ==0:
            Lm_day_plan = -100.0
        else:
            Lm_day_plan = L_sek_plan + 10 * math.log((passingtime_plan * trains_day_plan / 57600), 10)
        if trains_night_0 == 0:
            Lm_night_0 = -100.0
        else:
            Lm_night_0 = L_sek_0 + 10 * math.log((passingtime_0 * trains_night_0 / 28800), 10)
        if trains_night_plan ==0:
            Lm_night_plan = -100.0
        else:
            Lm_night_plan = L_sek_plan + 10 * math.log((passingtime_plan * trains_night_plan / 28800), 10)


    return Lm_day_0, Lm_night_0, Lm_day_plan, Lm_night_plan

def print2textview(Lm_day_0, Lm_night_0, Lm_day_plan, Lm_night_plan, check4150_string_0, check4150_string_plan,
                   air_check_0, air_check_plan, sig_KBF_max, sig_KBFtm, sig_KBFtr_day, sig_KBFtr_night, sig_Lm_day, sig_Lm_night):
    textview1 = builder.get_object('textview1')             #KBFmax Prognose-0
    textview2 = builder.get_object('textview2')             #KBFmax Prognose-Plan
    textview3 = builder.get_object('textview3')             #KBFtm Prognose-0
    textview4 = builder.get_object('textview4')             #KBFtm Prognose-Plan
    textview5 = builder.get_object('textview5')             #KBFtr tags Prognose-0
    textview6 = builder.get_object('textview6')             #KBFtr tags Prognose-Plan
    textview7 = builder.get_object('textview7')             #KBFtr nacht Prognose-0
    textview8 = builder.get_object('textview8')             #KBFtr nachts Prognose-Plan
    textview9 = builder.get_object('textview9')             #Lm tags Prognose-0
    textview10 = builder.get_object('textview10')           #Lm tags Prognose-Plan
    textview11 = builder.get_object('textview11')           #Lm nachts Prognose-0
    textview12 = builder.get_object('textview12')           #Lm nachts Prognose-Plan
    textview13 = builder.get_object('textview13')           #DIN 4150 Prognose-0
    textview14 = builder.get_object('textview14')           #DIN 4150 Prognose-Plan
    textview15 = builder.get_object('textview15')           #24. BImSchV Prognose-0
    textview16 = builder.get_object('textview16')           #24. BImSchV Prognose-plan
    textview18 = builder.get_object('textview18')           #Signifikanz KBFmax
    textview19 = builder.get_object('textview19')           #Signifikanz KBFtm
    textview20 = builder.get_object('textview20')           #Signifikanz KBFtr tags
    textview21 = builder.get_object('textview21')           #Signifikanz KBFtr nachts
    textview22 = builder.get_object('textview22')           #Signifikanz Lm tags
    textview23 = builder.get_object('textview23')           #Signifikanz Lm nachts

    ###### KBFmax 0 ######
    text_opt(textview1)
    buffer1 = textview1.get_buffer()
    buffer1.delete(buffer1.get_start_iter(), buffer1.get_end_iter())
    string1 = '%6.4f' % KBF_max_0
    buffer1.insert(buffer1.get_end_iter(), string1)
    ###### KBFmax plan ######
    text_opt(textview2)
    buffer2 = textview2.get_buffer()
    buffer2.delete(buffer2.get_start_iter(), buffer2.get_end_iter())
    string2 = '%6.4f' % KBF_max_plan
    buffer2.insert(buffer2.get_end_iter(), string2)
    ###### KBFtm 0 ###### textview 4
    text_opt(textview3)
    buffer3 = textview3.get_buffer()
    buffer3.delete(buffer3.get_start_iter(), buffer3.get_end_iter())
    string3 = '%6.4f' % KBFtm_0
    buffer3.insert(buffer3.get_end_iter(), string3)
    ###### KBFtm plan ######
    text_opt(textview4)
    buffer4 = textview4.get_buffer()
    buffer4.delete(buffer4.get_start_iter(), buffer4.get_end_iter())
    string4 = '%6.4f' % KBFtm_plan
    buffer4.insert(buffer4.get_end_iter(), string4)
    ###### KBFtr tags 0 ######
    text_opt(textview5)
    buffer5 = textview5.get_buffer()
    buffer5.delete(buffer5.get_start_iter(), buffer5.get_end_iter())
    string5 = '%6.4f' % KBFtr_day_0
    buffer5.insert(buffer5.get_end_iter(), string5)
    ###### KBFtr tags plan ######
    text_opt(textview6)
    buffer6 = textview6.get_buffer()
    buffer6.delete(buffer6.get_start_iter(), buffer6.get_end_iter())
    string6 = '%6.4f' % KBFtr_day_plan
    buffer6.insert(buffer6.get_end_iter(), string6)
    ###### KBFtr nachts 0 ######
    text_opt(textview7)
    buffer7 = textview7.get_buffer()
    buffer7.delete(buffer7.get_start_iter(), buffer7.get_end_iter())
    string7 = '%6.4f' % KBFtr_night_0
    buffer7.insert(buffer7.get_end_iter(), string7)
    ###### KBFtr nachts plan ######
    text_opt(textview8)
    buffer8 = textview8.get_buffer()
    buffer8.delete(buffer8.get_start_iter(), buffer8.get_end_iter())
    string8 = '%6.4f' % KBFtr_night_plan
    buffer8.insert(buffer8.get_end_iter(), string8)
    ###### Lm tags 0 ###########
    text_opt(textview9)
    buffer9 = textview9.get_buffer()
    buffer9.delete(buffer9.get_start_iter(), buffer9.get_end_iter())
    string9 = '%6.4f' % Lm_day_0
    buffer9.insert(buffer9.get_end_iter(), string9)
    ###### Lm tags plan ###########
    text_opt(textview10)
    buffer10 = textview10.get_buffer()
    buffer10.delete(buffer10.get_start_iter(), buffer10.get_end_iter())
    string10 = '%6.4f' % Lm_day_plan
    buffer10.insert(buffer10.get_end_iter(), string10)
    ###### Lm nachts 0 ###########
    text_opt(textview11)
    buffer11 = textview11.get_buffer()
    buffer11.delete(buffer11.get_start_iter(), buffer11.get_end_iter())
    string11 = '%6.4f' % Lm_night_0
    buffer11.insert(buffer11.get_end_iter(), string11)
    ###### Lm nachts plan ###########
    text_opt(textview12)
    buffer12 = textview12.get_buffer()
    buffer12.delete(buffer12.get_start_iter(), buffer12.get_end_iter())
    string12 = '%6.4f' % Lm_night_plan
    buffer12.insert(buffer12.get_end_iter(), string12)
    ###### DIN 4150 0 ###########
    text_opt(textview13)
    buffer13 = textview13.get_buffer()
    buffer13.delete(buffer13.get_start_iter(), buffer13.get_end_iter())
    string13 = '%s' % check4150_string_0
    buffer13.insert(buffer13.get_end_iter(), string13)
    ###### DIN 4150 plan ###########
    text_opt(textview14)
    buffer14 = textview14.get_buffer()
    buffer14.delete(buffer14.get_start_iter(), buffer14.get_end_iter())
    string14 = '%s' % check4150_string_plan
    buffer14.insert(buffer14.get_end_iter(), string14)
    ###### 24.BImSchV 0 ###########
    text_opt(textview15)
    buffer15 = textview15.get_buffer()
    buffer15.delete(buffer15.get_start_iter(), buffer15.get_end_iter())
    string15 = '%s' % air_check_0
    buffer15.insert(buffer15.get_end_iter(), string15)
    ###### 24.BImSchV plan ###########
    text_opt(textview16)
    buffer16 = textview16.get_buffer()
    buffer16.delete(buffer16.get_start_iter(), buffer16.get_end_iter())
    string16 = '%s' % air_check_plan
    buffer16.insert(buffer16.get_end_iter(), string16)

    ###### significance KBF max ###########
    text_opt(textview18)
    buffer18 = textview18.get_buffer()
    buffer18.delete(buffer18.get_start_iter(), buffer18.get_end_iter())
    string18 = '%4.2f%%' % sig_KBF_max
    buffer18.insert(buffer18.get_end_iter(), string18)
    ###### significance KBFtm ###########
    text_opt(textview19)
    buffer19 = textview19.get_buffer()
    buffer19.delete(buffer19.get_start_iter(), buffer19.get_end_iter())
    string19 = '%4.2f%%' % sig_KBFtm
    buffer19.insert(buffer19.get_end_iter(), string19)
    ###### significance KBFtr day ###########
    text_opt(textview20)
    buffer20 = textview20.get_buffer()
    buffer20.delete(buffer20.get_start_iter(), buffer20.get_end_iter())
    string20 = '%4.2f%%' % sig_KBFtr_day
    buffer20.insert(buffer20.get_end_iter(), string20)
    ###### significance KBFtr night ###########
    text_opt(textview21)
    buffer21 = textview21.get_buffer()
    buffer21.delete(buffer21.get_start_iter(), buffer21.get_end_iter())
    string21 = '%4.2f%%' % sig_KBFtr_night
    buffer21.insert(buffer21.get_end_iter(), string21)
    ###### significance Lm day ###########
    text_opt(textview22)
    buffer22 = textview22.get_buffer()
    buffer22.delete(buffer22.get_start_iter(), buffer22.get_end_iter())
    string22 = '%4.2f dB' % sig_Lm_day
    buffer22.insert(buffer22.get_end_iter(), string22)
    ###### significance Lm night ###########
    text_opt(textview23)
    buffer23 = textview23.get_buffer()
    buffer23.delete(buffer23.get_start_iter(), buffer23.get_end_iter())
    string23 = '%4.2f dB' % sig_Lm_night
    buffer23.insert(buffer23.get_end_iter(), string23)

def print_all_freq2textview(all_KBFtm_0, all_KBFtm_plan, all_KBF_max_0, all_KBF_max_plan, all_KBFtr_day_0, all_KBFtr_day_plan, all_KBFtr_night_0, all_KBFtr_night_plan, all_Lm_day_0, all_Lm_day_plan, all_Lm_night_0, all_Lm_night_plan, all_check4150_string_0, all_check4150_string_plan, all_air_check_0, all_air_check_plan, all_sig_KBF_max, all_sig_KBFtm, all_sig_KBFtr_day, all_sig_KBFtr_night, all_sig_Lm_day, all_sig_Lm_night):
    results_list = []
    results_list.append(all_KBF_max_0)
    results_list.append(all_KBF_max_plan)
    results_list.append(all_sig_KBF_max)
    results_list.append(all_KBFtm_0)
    results_list.append(all_KBFtm_plan)
    results_list.append(all_sig_KBFtm)
    results_list.append(all_KBFtr_day_0)
    results_list.append(all_KBFtr_day_plan)
    results_list.append(all_sig_KBFtr_day)
    results_list.append(all_KBFtr_night_0)
    results_list.append(all_KBFtr_night_plan)
    results_list.append(all_sig_KBFtr_night)
    results_list.append(all_Lm_day_0)
    results_list.append(all_Lm_day_plan)
    results_list.append(all_sig_Lm_day)
    results_list.append(all_Lm_night_0)
    results_list.append(all_Lm_night_plan)
    results_list.append(all_sig_Lm_night)
    results_list.append(all_check4150_string_0)
    results_list.append(all_check4150_string_plan)
    results_list.append(all_air_check_0)
    results_list.append(all_air_check_plan)

    for i in range(0, len(results_list[0])):
        for j in range(0, len(results_list)):
            text = all_frequencies.get_object(('text_' + str(i) + '_' + str(j)))
            text_opt_frqwin(text)
            buffer1 = text.get_buffer()
            buffer1.delete(buffer1.get_start_iter(), buffer1.get_end_iter())
            if j<=1 or j>2 and j<=4 or j>5 and j<=7 or j>8 and j<=10 or j>11 and j<=13 or j>14 and j<=16:
                string1 = '%6.4f' % results_list[j][i]
                buffer1.insert(buffer1.get_end_iter(), string1)
            elif j==2 or j==5 or j==8 or j==11 or j==14 or j==17:
                string1 = '%4.1f' % results_list[j][i]
                buffer1.insert(buffer1.get_end_iter(), string1)
            elif j > 17:
                string1 = '%s' % results_list[j][i]
                buffer1.insert(buffer1.get_end_iter(), string1)

def text_opt(textview):
    textview.set_property('editable', False)
    textview.set_justification(Gtk.Justification.RIGHT)

def text_opt_frqwin(textview):
        textview.set_property('editable', False)
        textview.set_justification(Gtk.Justification.CENTER)

def initialize_builder():
    builder = Gtk.Builder()
    builder.add_objects_from_file('./glade/glade_gui_gen1.06.glade', ('analysis_window', '') )
    builder.add_objects_from_file('./glade/glade_gui_gen1.06.glade', ('configure_window', ''))
    builder.add_objects_from_file('./glade/glade_gui_gen1.06.glade', ('location_dialog', ''))
    builder.add_objects_from_file('./glade/glade_gui_gen1.06.glade', ('main_window', ''))
    builder.add_objects_from_file('./glade/glade_gui_gen1.06.glade', ('results_single', ''))
    builder.add_objects_from_file('./glade/glade_gui_gen1.06.glade', ('plot_universal', ''))
    builder.add_objects_from_file('./glade/glade_gui_gen1.06.glade', ('prognose_data_single', ''))
    builder.add_objects_from_file('./glade/glade_gui_gen1.06.glade', ('switch_dialog', ''))
    builder.connect_signals(Signals())
    return builder

def initialize_custom_reductions_window():
    custom_reductions = Gtk.Builder()
    custom_reductions.add_objects_from_file('./glade/glade_custom_window.glade', ('custom_window', ''))
    custom_reductions.connect_signals(Signals())
    return custom_reductions

def initialize_view_all_frequencies_window():
    all_frequencies = Gtk.Builder()
    all_frequencies.add_objects_from_file('./glade/glade_all_frequencies_window1.06.glade', ('all_frequencies_window', ''))
    all_frequencies.connect_signals(Signals())
    return all_frequencies

def initialize_scrolledwindow1(builder):
    sw1 = builder.get_object('scrolledwindow1')
    fig1 = Figure(figsize=(4,4), dpi=100)
    ax1 = fig1.add_subplot(111)
    common_ax1(ax1)
    canvas = FigureCanvas(fig1)
    sw1.add_with_viewport(canvas)
    return sw1, fig1, ax1

def initialize_scrolledwindow2(builder):
    sw2 = builder.get_object('scrolledwindow2')
    fig2 = Figure(figsize=(4,4), dpi=100)
    ax2 = fig2.add_subplot(111)
    common_ax2(ax2)
    canvas = FigureCanvas(fig2)
    sw2.add_with_viewport(canvas)
    return sw2, fig2, ax2

def initialize_scrolledwindow3(builder):
    sw3 = builder.get_object('scrolledwindow3')
    fig3 = Figure(figsize=(4,4), dpi=100)
    ax3 = fig3.add_subplot(111)
    common_ax3(ax3)
    canvas = FigureCanvas(fig3)
    sw3.add_with_viewport(canvas)
    return sw3, fig3, ax3

def initialize_scrolledwindow4(builder):
    sw4 = builder.get_object('scrolledwindow4')
    fig4 = Figure(figsize=(4, 3), dpi=100)
    ax4 = fig4.add_subplot(111)
    common_ax4(ax4)
    fig4.tight_layout()
    canvas = FigureCanvas(fig4)
    sw4.add_with_viewport(canvas)
    return sw4, fig4, ax4

def initialize_scrolledwindow5(builder):
    sw5 = builder.get_object('scrolledwindow5')
    fig5 = Figure(figsize=(4,3), dpi=100)
    ax5 = fig5.add_subplot(111)
    common_ax5(ax5)
    fig5.tight_layout()
    canvas = FigureCanvas(fig5)
    sw5.add_with_viewport(canvas)
    return sw5, fig5, ax5

def initialize_scrolledwindow6(builder):
    sw6 = builder.get_object('scrolledwindow6')
    fig6 = Figure(figsize=(4,4), dpi=100)
    ax6 = fig6.add_subplot(111)
    common_ax6(ax6)
    canvas = FigureCanvas(fig6)
    sw6.add_with_viewport(canvas)
    return sw6, fig6, ax6

def initialize_scrolledwindow7(builder):
    sw7 = builder.get_object('scrolledwindow_all_results')
    fig7 = Figure(figsize=(6,6), dpi=100)
    #fig1.tight_layout()
    ax7 = fig7.add_subplot(111)
    common_ax7(ax7)
    canvas = FigureCanvas(fig7)
    sw7.add_with_viewport(canvas)
    return sw7, fig7, ax7

def initialize_scrolledwindow8(builder):
    sw8 = builder.get_object('scrolledwindow_universal')
    fig8 = Figure(figsize=(6,6), dpi=100)
    #fig1.tight_layout()
    ax8 = fig8.add_subplot(111)
    common_ax8(ax8)
    canvas = FigureCanvas(fig8)
    sw8.add_with_viewport(canvas)
    return sw8, fig8, ax8

def initialize_configure_window_objects(builder):
    textview_location_preview = builder.get_object('textview_location_preview')
    entry_notes = builder.get_object('entry_notes')
    combobox_landuse = builder.get_object('combobox_landuse')
    combobox_air_category = builder.get_object('combobox_category')
    return textview_location_preview, entry_notes, combobox_landuse,combobox_air_category

def initialize_location_dialog_objects(builder):
    entry_track = builder.get_object('entry_track')
    entry_track_loc = builder.get_object('entry_track_loc')
    entry_street = builder.get_object('entry_street')
    entry_postal = builder.get_object('entry_postal')
    entry_location = builder.get_object('entry_location')
    combobox_state = builder.get_object('combobox_state')
    combobox_country = builder.get_object('combobox_country')
    return entry_track, entry_track_loc, entry_street, entry_postal, entry_location, combobox_state, combobox_country

def initialize_main_window_objects(builder):
    #buttons
    button_apply1 = builder.get_object('button1')
    button_apply2 = builder.get_object('button2')
    button_linear = builder.get_object('button3')
    button_Hoelzl = builder.get_object('button6')
    button_linear_model = builder.get_object('button_linear_model')
    button_Transferfunction_from_file = builder.get_object('button7')
    button_custom_tfbld = builder.get_object('button17')
    button_switch = builder.get_object('button11')
    button_modified_transferfunction = builder.get_object('button12')
    button_apply3 = builder.get_object('button13')
    button_write_variables2results = builder.get_object('button_write_variables2results')
    #infobutton
    button_info = builder.get_object('button8')
    image_i = Gtk.Image()
    image_i.set_from_file('./gfx/i_10px.png')
    button_info.set_image(image_i)

    #textview Fehler und Warnungen
    textview17 = builder.get_object('textview17')

    #menuitems
    menu_file_open = builder.get_object('menuitem_file_open')
    menu_file_from_database = builder.get_object('menuitem_file_from_database')
    # comboboxes
    combobox_frequency = builder.get_object('comboboxtext1')
    combobox_ceiling = builder.get_object('comboboxtext2')
    combobox_dataset4analysis =builder.get_object('combobox_dataset4analysis')

    return button_apply1, button_apply2, button_linear, button_Hoelzl, button_linear_model, button_Transferfunction_from_file, button_custom_tfbld, textview17, \
           button_switch, button_modified_transferfunction, button_apply3, combobox_frequency, combobox_ceiling, menu_file_open,\
           menu_file_from_database, combobox_dataset4analysis, button_write_variables2results

def initialize_result_window_buttons(builder):
    button_traintype = []
    for i in range(0,30):
        button_traintype.append(builder.get_object('train_'+str(i)))
        #button_del_traintype.append(builder.get_object('del_'+str(i)))
    return button_traintype#, button_del_traintype

def interact1_set_active():
    builder.get_object('entry_MP5').set_sensitive(True)
    builder.get_object('entry_MP4').set_sensitive(True)
    builder.get_object('entry_dist_bld').set_sensitive(True)
    builder.get_object('entry_track_shift').set_sensitive(True)
    button_apply1.set_sensitive(True)

def interact2_set_active():
    if analysis_mode == 'zero':
        builder.get_object('entry_vel_measure').set_sensitive(True)
        builder.get_object('entry_vel_zero').set_sensitive(True)
    elif analysis_mode == 'plan':
        builder.get_object('entry_vel_measure').set_sensitive(True)
        builder.get_object('entry_vel_plan').set_sensitive(True)
    else:
        builder.get_object('entry_vel_measure').set_sensitive(True)
        builder.get_object('entry_vel_zero').set_sensitive(True)
        builder.get_object('entry_vel_plan').set_sensitive(True)
    button_apply2.set_sensitive(True)

def interact3_set_active():
    button_linear.set_sensitive(True)
    #DBRichtlinie.set_sensitive(True)
    #button_inversion.set_sensitive(True)
    button_Hoelzl.set_sensitive(True)
    if model == True:
        button_linear_model.set_sensitive(True)

def interact4_set_active():
    builder.get_object('entry_trains_day_zero').set_sensitive(True)
    builder.get_object('entry_trains_night_zero').set_sensitive(True)
    builder.get_object('entry_trains_day_plan').set_sensitive(True)
    builder.get_object('entry_trains_night_plan').set_sensitive(True)
    builder.get_object('entry_trainlength_zero').set_sensitive(True)
    builder.get_object('entry_trainlength_plan').set_sensitive(True)
    combobox_landuse.set_sensitive(True)
    combobox_air_category.set_sensitive(True)
    button_apply3.set_sensitive(True)

def set_transferfunction_combobox_active():
    combobox_ceiling.set_sensitive(True)
    '''
    if type(transferfunction) != None:# and mode != 'norm':
        combobox_ceiling.set_sensitive(False)
        plot_transferfunction(transferfunction, 'aus Messung', ax4, ticks)
        sw4.queue_draw()
        interact4_set_active()
        plot_reductions(custom_function, reduction_type, ax5, ticks)
        sw5.queue_draw()
    else:
        combobox_ceiling.set_sensitive(True)
    '''
def set_transferfunction_interact_active():
    if datalength == 5:
        combobox_frequency.set_sensitive(True)
        button_Transferfunction_from_file.set_sensitive(True)
        button_custom_tfbld.set_sensitive(True)
    else:
        combobox_frequency.set_sensitive(True)
        button_custom_tfbld.set_sensitive(True)

def set_reduction_interact_active():
    button_switch.set_sensitive(True)
    button_modified_transferfunction.set_sensitive(True)
    button_mind1.set_sensitive(True)
    button_mind2.set_sensitive(True)

def set_entries_passive():
    builder.get_object('entry_MP5').set_sensitive(False)
    builder.get_object('entry_MP4').set_sensitive(False)
    builder.get_object('entry_dist_bld').set_sensitive(False)
    builder.get_object('entry_track_shift').set_sensitive(False)
    builder.get_object('entry_vel_measure').set_sensitive(False)
    builder.get_object('entry_vel_zero').set_sensitive(False)
    builder.get_object('entry_vel_plan').set_sensitive(False)
    builder.get_object('entry_trains_day_zero').set_sensitive(False)
    builder.get_object('entry_trains_night_zero').set_sensitive(False)
    builder.get_object('entry_trains_day_plan').set_sensitive(False)
    builder.get_object('entry_trains_night_plan').set_sensitive(False)
    builder.get_object('entry_trainlength_zero').set_sensitive(False)
    builder.get_object('entry_trainlength_plan').set_sensitive(False)

def set_buttons_passive():
    button_apply1.set_sensitive(False)
    button_apply2.set_sensitive(False)
    button_linear.set_sensitive(False)
    button_Hoelzl.set_sensitive(False)
    button_linear_model.set_sensitive(False)
    button_Transferfunction_from_file.set_sensitive(False)
    button_custom_tfbld.set_sensitive(False)
    button_switch.set_sensitive(False)
    button_modified_transferfunction.set_sensitive(False)
    button_apply3.set_sensitive(False)
    button_write_variables2results.set_sensitive(False)

def set_menuitems_passive():
    menu_file_open.set_sensitive(False)
    menu_file_from_database.set_sensitive(False)
    builder.get_object('Gebäudeankopplung').set_sensitive(False)
    builder.get_object('Deckenübertragung').set_sensitive(False)

def set_menuitems_active():
    menu_file_open.set_sensitive(True)
    menu_file_from_database.set_sensitive(True)

def set_comboboxes_passive():
    combobox_frequency.set_sensitive(False)
    combobox_ceiling.set_sensitive(False)
    #combobox_dataset4analysis.set_sensitive(False)

def set_textviews_passive():
    textview1.set_sensitive(False)
    textview2.set_sensitive(False)
    textview3.set_sensitive(False)
    textview4.set_sensitive(False)
    textview5.set_sensitive(False)
    textview6.set_sensitive(False)
    textview7.set_sensitive(False)
    textview8.set_sensitive(False)

#def set_add_to_results_active():
#    combobox_dataset4analysis.set_sensitive(True)

def check_4150_conditions(KBF_max_0, KBF_max_plan, KBFtr_day_0, KBFtr_night_0, KBFtr_day_plan, KBFtr_night_plan, Au_day, Au_night, Ao_day, Ao_night, Ar_day, Ar_night):
    # modified 170125 deleted Ao criteria because of no use for railway calculations
    global warnings
    warnings = []
    check4150_string_0 = str
    check4150_string_plan = str

    if builder.get_object('switch1').get_state() == True:
        Au_day = Au_day*1.5
        Au_night = Au_night*1.5
        Ar_day = Ar_day*1.5
        Ar_night = Ar_night*1.5

    if KBF_max_0 <= Au_day and KBF_max_0 <= Au_night:
        check4150_string_0 = "eingehalten"
    elif KBF_max_0 > Au_day or KBF_max_0 > Au_night:
        if KBFtr_day_0 <= Ar_day and KBFtr_night_0 <= Ar_night:
            check4150_string_0 = "eingehalten"
        elif KBFtr_day_0 <= Ar_day and KBFtr_night_0 > Ar_night:
            check4150_string_0 = "nicht eingehalten"
            warnings.append("KBFtr nachts 0-Fall > Anhaltswert Ar nachts")
        elif KBFtr_day_0 > Ar_day and KBFtr_night_0 <= Ar_night:
            check4150_string_0 = "nicht eingehalten"
            warnings.append("KBFtr tags 0-Fall > Anhaltswert Ar tags")
        else:
            check4150_string_0 = "nicht eingehalten"
            warnings.append("KBFtr tags Nullfall > Ar tags")
            warnings.append("KBFtr nachts Nullfall > Ar nachts")

    if KBF_max_plan <= Au_day and KBF_max_plan <= Au_night:
        check4150_string_plan = "eingehalten"
    elif KBF_max_plan > Au_day or KBF_max_plan > Au_night:
        if KBFtr_day_plan <= Ar_day and KBFtr_night_plan <= Ar_night:
            check4150_string_plan = "eingehalten"
        elif KBFtr_day_plan <= Ar_day and KBFtr_night_plan > Ar_night:
            check4150_string_plan = "nicht eingehalten"
            warnings.append("KBFtr nachts Plan-Fall > Anhaltswert Ar nachts")
        elif KBFtr_day_plan > Ar_day and KBFtr_night_plan <= Ar_night:
            check4150_string_plan = "nicht eingehalten"
            warnings.append("KBFtr tags Plan-Fall > Anhaltswert Ar tags")
        else:
            check4150_string_plan = "nicht eingehalten"
            warnings.append("KBFtr tags Planfall > Ar tags")
            warnings.append("KBFtr nachts Planfall > Ar nachts")

    return check4150_string_0, check4150_string_plan

def check_sec_airborne_sound(air_category, Lm_day_0, Lm_night_0, Lm_day_plan, Lm_night_plan):
    air_day, air_night = read_category_values_from_table(air_category)
    if air_category == 'Kat. 1':
        if Lm_day_0 <= air_day and Lm_night_0 <= air_night:
            air_check_0 = 'eingehalten'
        elif Lm_day_0 > air_day and Lm_night_0 <= air_night:
            air_check_0 = 'nicht eingehalten'
            warnings.append('Lm 0-Fall > Anhaltswert Lm tags')
        elif Lm_day_0 <= air_day and Lm_night_0 > air_night:
            air_check_0 = 'nicht eingehalten'
            warnings.append('Lm 0-Fall > Anhaltswert Lm nachts')
        else:
            air_check_0 = 'nicht eingehalten'
            warnings.append('Lm 0-Fall > Anhaltswert Lm tags')
            warnings.append('Lm 0-Fall > Anhaltswert Lm nachts')
    elif air_category == 'Kat. 2':
        if Lm_night_0 <= air_night:
            air_check_0 = 'eingehalten'
        else:
            air_check_0 = 'nicht eingehalten'
            warnings.append('Lm 0-Fall > Anhaltswert Lm nachts')
    elif air_category == 'Kat. 3':
        if Lm_day_0 <= air_day:
            air_check_0 = 'eingehalten'
        else:
            air_check_0 = 'nicht eingehalten'
            warnings.append('Lm 0-Fall > Anhaltswert Lm tags')
    elif air_category == 'Kat. 4':
        if Lm_day_0 <= air_day:
            air_check_0 = 'eingehalten'
        else:
            air_check_0 = 'nicht eingehalten'
            warnings.append('Lm 0-Fall > Anhaltswert Lm tags')
    elif air_category == 'Kat. 5':
        if Lm_day_0 <= air_day:
            air_check_0 = 'eingehalten'
        else:
            air_check_0 = 'nicht eingehalten'
            warnings.append('Lm 0-Fall > Anhaltswert Lm tags')
    elif air_category == 'Kat. 6':
        if Lm_day_0 <= air_day:
            air_check_0 = 'eingehalten'
        else:
            air_check_0 = 'nicht eingehalten'
            warnings.append('Lm 0-Fall > Anhaltswert Lm tags')

    if air_category == 'Kat. 1':
        if Lm_day_plan <= air_day and Lm_night_plan <= air_night:
            air_check_plan = 'eingehalten'
        elif Lm_day_plan > air_day and Lm_night_plan <= air_night:
            air_check_plan = 'nicht eingehalten'
            warnings.append('Lm Plan-Fall > Anhaltswert Lm tags')
        elif Lm_day_plan <= air_day and Lm_night_plan > air_night:
            air_check_plan = 'nicht eingehalten'
            warnings.append('Lm Plan-Fall > Anhaltswert Lm nachts')
        else:
            air_check_plan = 'nicht eingehalten'
            warnings.append('Lm Plan-Fall > Anhaltswert Lm tags')
            warnings.append('Lm Plan-Fall > Anhaltswert Lm nachts')
    elif air_category == 'Kat. 2':
        if Lm_night_plan <= air_night:
            air_check_plan = 'eingehalten'
        else:
            air_check_plan = 'nicht eingehalten'
            warnings.append('Lm Plan-Fall > Anhaltswert Lm nachts')
    elif air_category == 'Kat. 3':
        if Lm_day_plan <= air_day:
            air_check_plan = 'eingehalten'
        else:
            air_check_plan = 'nicht eingehalten'
            warnings.append('Lm Plan-Fall > Anhaltswert Lm tags')
    elif air_category == 'Kat. 4':
        if Lm_day_plan <= air_day:
            air_check_plan = 'eingehalten'
        else:
            air_check_plan = 'nicht eingehalten'
            warnings.append('Lm Plan-Fall > Anhaltswert Lm tags')
    elif air_category == 'Kat. 5':
        if Lm_day_plan <= air_day:
            air_check_plan = 'eingehalten'
        else:
            air_check_plan = 'nicht eingehalten'
            warnings.append('Lm Plan-Fall > Anhaltswert Lm tags')
    elif air_category == 'Kat. 6':
        if Lm_day_plan <= air_day:
            air_check_plan = 'eingehalten'
        else:
            air_check_plan = 'nicht eingehalten'
            warnings.append('Lm Plan-Fall > Anhaltswert Lm tags')

    return air_check_0, air_check_plan

def warnings2textview(warnings):
    text_opt(textview17)
    buffer17 = textview17.get_buffer()
    buffer17.delete(buffer17.get_start_iter(), buffer17.get_end_iter())
    warnstring = ''
    if warning_vel != 0:
        warnstring = warning_vel + '\n'
    for i in range(0, len(warnings)):
        warnstring_temp = '%s' % str(warnings[i])
        warnstring = str(warnstring) + str(warnstring_temp) + '\n'
    warnings = []
    buffer17.insert(buffer17.get_end_iter(), warnstring)

def calc_significance(KBF_max_0, KBF_max_plan, KBFtm_0, KBFtm_plan, KBFtr_day_0, KBFtr_night_0, KBFtr_day_plan, KBFtr_night_plan, Lm_day_0, Lm_night_0, Lm_day_plan, Lm_night_plan):
    sig_KBF_max = 100*KBF_max_plan/KBF_max_0-100
    sig_KBFtm = 100*KBFtm_plan/KBFtm_0-100
    if KBFtr_day_0 ==0 or KBFtr_night_0 ==0:
        sig_KBFtr_day = np.nan
    else:
        sig_KBFtr_day = 100*KBFtr_day_plan/KBFtr_day_0-100
    if KBFtr_night_0 == 0 or KBFtr_night_plan == 0:
        sig_KBFtr_night = np.nan
    else:
        sig_KBFtr_night = 100*KBFtr_night_plan/KBFtr_night_0-100
    sig_Lm_day = Lm_day_plan-Lm_day_0
    sig_Lm_night = Lm_night_plan-Lm_night_0
    return sig_KBF_max, sig_KBFtm, sig_KBFtr_day, sig_KBFtr_night, sig_Lm_day, sig_Lm_night

def calc_all(Lv_outsideBld, Lv_outsideBld_prospective, Lv_outsideBld_Leq, Lv_outsideBld_prospective_Leq, custom_function,terzfrequencies, trains_day_0, trains_night_0,
             trains_day_plan, trains_night_plan, ceiling, _0, trainlength_plan, vel_measure, vel_prognose, Au_day, Au_night, Ao_day,
                                                                          Ao_night, Ar_day, Ar_night, air_category):
    global all_frequencies_calculation
    all_frequencies_calculation = {}

    all_transferfunctions=[]
    all_level_inside_0 = []
    all_level_inside_0_Leq = []
    all_ceiling = []
    all_level_inside_plan = []
    all_level_inside_plan_Leq = []
    all_v_inside_0 = []
    all_v_inside_plan = []
    all_KBF_0 = []
    all_KBF_plan = []
    all_KBFtm_0 = []
    all_KBFtm_plan = []
    all_KBF_max_0 = []
    all_KBF_max_plan = []
    all_KBFtr_day_0 = []
    all_KBFtr_day_plan = []
    all_KBFtr_night_0 = []
    all_KBFtr_night_plan = []
    all_Lm_day_0 = []
    all_Lm_night_0 = []
    all_Lm_day_plan = []
    all_Lm_night_plan = []
    all_check4150_string_0 = []
    all_check4150_string_plan = []
    all_air_check_0 = []
    all_air_check_plan = []
    all_sig_KBF_max = []
    all_sig_KBFtm = []
    all_sig_KBFtr_day =[]
    all_sig_KBFtr_night = []
    all_sig_Lm_day = []
    all_sig_Lm_night = []
    all_freq = ['8 Hz', '10 Hz','12.5 Hz','16 Hz','20 Hz','25 Hz','31.5 Hz','40 Hz','50 Hz','62.5 Hz','80 Hz']
    for item in all_freq:
        all_transferfunctions.append(read_transferfunctions_from_table(item,'./src/transferfunctions_wood'))
        all_ceiling.append('Holz')
    for item in all_freq:
        all_transferfunctions.append(read_transferfunctions_from_table(item, './src/transferfunctions_concrete'))
        all_ceiling.append('Beton')
    for item in all_transferfunctions:
        temp_all_level_inside_0, temp_all_level_inside_plan, temp_all_level_inside_0_Leq, temp_all_level_inside_plan_Leq =calc_level_inside(Lv_outsideBld, Lv_outsideBld_prospective, Lv_outsideBld_Leq, Lv_outsideBld_prospective_Leq, item, custom_function)
        if analysis_mode == 'zero':
            all_level_inside_0.append(temp_all_level_inside_0)
            all_level_inside_0_Leq.append(temp_all_level_inside_0_Leq)
            all_level_inside_plan.append(np.zeros(20))
            all_level_inside_plan_Leq.append(np.zeros(20))
        elif analysis_mode == 'plan':
            all_level_inside_0.append(np.zeros(20))
            all_level_inside_0_Leq.append(np.zeros(20))
            all_level_inside_plan.append(temp_all_level_inside_plan)
            all_level_inside_plan_Leq.append(temp_all_level_inside_plan_Leq)
        else:
            all_level_inside_0.append(temp_all_level_inside_0)
            all_level_inside_0_Leq.append(temp_all_level_inside_0_Leq)
            all_level_inside_plan.append(temp_all_level_inside_plan)
            all_level_inside_plan_Leq.append(temp_all_level_inside_plan_Leq)
    for i in range(0,len(all_level_inside_0)):
        temp_v_inside_0, temp_v_inside_plan = calc_vel_inside(all_level_inside_0[i], all_level_inside_plan[i])
        temp_KBF_0, temp_KBF_plan = calc_KBF(temp_v_inside_0, temp_v_inside_plan, terzfrequencies)
        temp_KBFtm_0, temp_KBFtm_plan = calc_KBFtm(temp_KBF_0, temp_KBF_plan)
        temp_KBF_max_0, temp_KBF_max_plan = calc_KBFmax(temp_KBFtm_0, temp_KBFtm_plan, all_ceiling[i])
        temp_KBFtr_day_0, temp_KBFtr_night_0, temp_KBFtr_day_plan, temp_KBFtr_night_plan = calc_KBFtr(temp_KBFtm_0, temp_KBFtm_plan, trains_day_0,
                                                                                  trains_night_0, trains_day_plan,
                                                                                  trains_night_plan)
        temp_Lm_day_0, temp_Lm_night_0, temp_Lm_day_plan, temp_Lm_night_plan = calc_secondary_airborne_sound(all_level_inside_0_Leq[i], all_level_inside_plan_Leq[i], ceiling, trainlength_0, trainlength_plan, vel_measure, vel_prognose,
                                                                                         trains_day_0, trains_night_0, trains_day_plan,
                                                                                         trains_night_plan)

        temp_check4150_string_0, temp_check4150_string_plan = check_4150_conditions(temp_KBF_max_0, temp_KBF_max_plan, temp_KBFtr_day_0,
                                                                          temp_KBFtr_night_0, temp_KBFtr_day_plan,
                                                                          temp_KBFtr_night_plan, Au_day, Au_night, Ao_day,
                                                                          Ao_night, Ar_day, Ar_night)

        temp_air_check_0, temp_air_check_plan = check_sec_airborne_sound(air_category, temp_Lm_day_0, temp_Lm_night_0, temp_Lm_day_plan,
                                                               temp_Lm_night_plan)
        temp_sig_KBF_max,  temp_sig_KBFtm,  temp_sig_KBFtr_day,  temp_sig_KBFtr_night,  temp_sig_Lm_day,  temp_sig_Lm_night = calc_significance(temp_KBF_max_0, temp_KBF_max_plan,
                                                                                                             temp_KBFtm_0, temp_KBFtm_plan,
                                                                                                             temp_KBFtr_day_0,temp_KBFtr_night_0,
                                                                                                             temp_KBFtr_day_plan, temp_KBFtr_night_plan,
                                                                                                             temp_Lm_day_0, temp_Lm_night_0,temp_Lm_day_plan,
                                                                                                             temp_Lm_night_plan)
        if analysis_mode == 'zero':
            all_v_inside_0.append(temp_v_inside_0)
            all_KBF_0.append(temp_KBF_0)
            all_KBFtm_0.append(temp_KBFtm_0)
            all_KBF_max_0.append(temp_KBF_max_0)
            all_KBFtr_day_0.append(temp_KBFtr_day_0)
            all_KBFtr_night_0.append(temp_KBFtr_night_0)
            all_Lm_day_0.append(temp_Lm_day_0)
            all_Lm_night_0.append(temp_Lm_night_0)
            all_check4150_string_0.append(temp_check4150_string_0)
            all_air_check_0.append(temp_air_check_0)

            all_v_inside_plan.append(np.zeros(20))
            all_KBF_plan.append(0.0)
            all_KBFtm_plan.append(0.0)
            all_KBF_max_plan.append(0.0)
            all_KBFtr_day_plan.append(0.0)
            all_KBFtr_night_plan.append(0.0)
            all_Lm_day_plan.append(-100.0)
            all_Lm_night_plan.append(-100.0)
            all_check4150_string_plan.append('-')
            all_air_check_plan.append('-')

            all_sig_KBF_max.append(0.0)
            all_sig_KBFtm.append(0.0)
            all_sig_KBFtr_day.append(0.0)
            all_sig_KBFtr_night.append(0.0)
            all_sig_Lm_day.append(0.0)
            all_sig_Lm_night.append(0.0)

        elif analysis_mode == 'plan':
            all_v_inside_0.append(np.zeros(20))
            all_KBF_0.append(0.0)
            all_KBFtm_0.append(0.0)
            all_KBF_max_0.append(0.0)
            all_KBFtr_day_0.append(0.0)
            all_KBFtr_night_0.append(0.0)
            all_Lm_day_0.append(-100.0)
            all_Lm_night_0.append(-100.0)
            all_check4150_string_0.append('-')
            all_air_check_0.append('-')

            all_v_inside_plan.append(temp_v_inside_plan)
            all_KBF_plan.append(temp_KBF_plan)
            all_KBFtm_plan.append(temp_KBFtm_plan)
            all_KBF_max_plan.append(temp_KBF_max_plan)
            all_KBFtr_day_plan.append(temp_KBFtr_day_plan)
            all_KBFtr_night_plan.append(temp_KBFtr_night_plan)
            all_Lm_day_plan.append(temp_Lm_day_plan)
            all_Lm_night_plan.append(temp_Lm_night_plan)
            all_check4150_string_plan.append(temp_check4150_string_plan)
            all_air_check_plan.append(temp_air_check_plan)

            all_sig_KBF_max.append(0.0)
            all_sig_KBFtm.append(0.0)
            all_sig_KBFtr_day.append(0.0)
            all_sig_KBFtr_night.append(0.0)
            all_sig_Lm_day.append(0.0)
            all_sig_Lm_night.append(0.0)
        else:
            all_v_inside_0.append(temp_v_inside_0)
            all_KBF_0.append(temp_KBF_0)
            all_KBFtm_0.append(temp_KBFtm_0)
            all_KBF_max_0.append(temp_KBF_max_0)
            all_KBFtr_day_0.append(temp_KBFtr_day_0)
            all_KBFtr_night_0.append(temp_KBFtr_night_0)
            all_Lm_day_0.append(temp_Lm_day_0)
            all_Lm_night_0.append(temp_Lm_night_0)
            all_check4150_string_0.append(temp_check4150_string_0)
            all_air_check_0.append(temp_air_check_0)

            all_v_inside_plan.append(temp_v_inside_plan)
            all_KBF_plan.append(temp_KBF_plan)
            all_KBFtm_plan.append(temp_KBFtm_plan)
            all_KBF_max_plan.append(temp_KBF_max_plan)
            all_KBFtr_day_plan.append(temp_KBFtr_day_plan)
            all_KBFtr_night_plan.append(temp_KBFtr_night_plan)
            all_Lm_day_plan.append(temp_Lm_day_plan)
            all_Lm_night_plan.append(temp_Lm_night_plan)
            all_check4150_string_plan.append(temp_check4150_string_plan)
            all_air_check_plan.append(temp_air_check_plan)

            all_sig_KBF_max.append(temp_sig_KBF_max)
            all_sig_KBFtm.append(temp_sig_KBFtm)
            all_sig_KBFtr_day.append(temp_sig_KBFtr_day)
            all_sig_KBFtr_night.append(temp_sig_KBFtr_night)
            all_sig_Lm_day.append(temp_sig_Lm_day)
            all_sig_Lm_night.append(temp_sig_Lm_night)

    all_frequencies_calculation ={'all_KBFtm_0':all_KBFtm_0, 'all_KBFtm_plan':all_KBFtm_plan, 'all_KBF_max_0':all_KBF_max_0,'all_KBF_max_plan':all_KBF_max_plan,
                                  'all_KBFtr_day_0':all_KBFtr_day_0, 'all_KBFtr_day_plan':all_KBFtr_day_plan, 'all_KBFtr_night_0': all_KBFtr_night_0,
                                  'all_KBFtr_night_plan':all_KBFtr_night_plan, 'all_Lm_day_0':all_Lm_day_0, 'all_Lm_day_plan': all_Lm_day_plan, 'all_Lm_night_0':all_Lm_night_0,
                                  'all_Lm_night_plan': all_Lm_night_plan, 'all_check4150_string_0':all_check4150_string_0, 'all_check4150_string_plan':all_check4150_string_plan,
                                  'all_air_check_0': all_air_check_0, 'all_air_check_plan':all_air_check_plan, 'all_sig_KBF_max':all_sig_KBF_max, 'all_sig_KBFtm':all_sig_KBFtm,
                                  'all_sig_KBFtr_day':all_sig_KBFtr_day, 'all_sig_KBFtr_night':all_sig_KBFtr_night, 'all_sig_Lm_day':all_sig_Lm_day, 'all_sig_Lm_night':all_sig_Lm_night}

    return all_KBFtm_0, all_KBFtm_plan, all_KBF_max_0, all_KBF_max_plan, all_KBFtr_day_0, all_KBFtr_day_plan, all_KBFtr_night_0, all_KBFtr_night_plan, all_Lm_day_0, all_Lm_day_plan, all_Lm_night_0, all_Lm_night_plan, all_check4150_string_0, all_check4150_string_plan, all_air_check_0, all_air_check_plan, all_sig_KBF_max, all_sig_KBFtm, all_sig_KBFtr_day, all_sig_KBFtr_night, all_sig_Lm_day, all_sig_Lm_night

def get_text_from_combobox(combobox):
    index = combobox.get_active()
    model = combobox.get_model()
    text = model[index][1]
    return text

def dataset_entries2configure_window(count, builder):
    global apply_button
    box = builder.get_object('box11')
    dataset_entry = []
    dataset_label = []
    window_height = [370,433,467,501,535,569,603,637,671,705,739,773,807,841,875,909,943,977,1011,1045]
    if len(box.get_children()) == 4:
        grid = box.get_children()[3]
        box.remove(grid)
        #print type(child[2])

    grid = Gtk.Grid()
    apply_button = Gtk.Button('übernehmen')

    for i in range(0, count):
        dataset_entry.append(Gtk.Entry())
        dataset_label.append(Gtk.Label("Bezeichnung Datensatz %s" % (i+1), xalign=1))
        grid.attach(dataset_label[i], 0, i, 1, 1)
        grid.attach_next_to(dataset_entry[i], dataset_label[i], Gtk.PositionType.RIGHT, 1, 1)
        dataset_entry[i].set_max_length(20)
        dataset_label[i].set_margin_left(20)
        dataset_label[i].set_margin_right(60)
        grid.set_row_spacing(2)
    grid.attach(apply_button, 1, (count+1), 1, 1 )
    signals = Signals()
    apply_button.connect("clicked", signals.apply_configuration)
    box.pack_start(grid, True, True, 0)
    configure_prognose.resize(439, window_height[count-1])
    grid.show_all()
    return sorted(dataset_entry)

'''
def fill_combobox_dataset4analysis(combobox, dataset_entries):
    Gtk.CellLayout.clear(combobox)
    liststore = Gtk.ListStore(int, str)
    elements = []
    for i in range(0, len(dataset_entries)):
        item = dataset_entries[i].get_text()
        liststore.append([i, str(item)])
        elements.append(item)
    combobox.set_model(liststore)
    cell = Gtk.CellRendererText()
    combobox.pack_start(cell, True)
    combobox.add_attribute(cell, 'text', 1)
    return elements
'''

def get_elements_from_entries(dataset_entries):
    elements = []
    for i in range(0, len(dataset_entries)):
        item = dataset_entries[i].get_text()
        elements.append(item)
    return elements

def adapt_analysis_results(elements, glob_notes):
    analysis_results['metadata'] = {'tracknumber':tracknumber, 'track_location':track_loc, 'street':street, 'postal':postal, 'location':location, 'state':state, 'country':country, 'notes':glob_notes}
    trainlist = analysis_results['analysis'].keys()
    if trainlist == []:
        for item in elements:
            analysis_results['analysis'][item] = {}

    #elif sorted(trainlist) == sorted(elements):
        #nothing to do

    elif sorted(trainlist) != sorted(elements):
        if len(trainlist) < len(elements):
            diff = list(set.difference(set(elements), set(trainlist)))
            for item in diff:
                analysis_results['analysis'][item] = {}

        elif len(trainlist) > len(elements):
            diff = list(set.difference(set(trainlist), set(elements)))
            for item in diff:
                del analysis_results['analysis'][item]

        elif len(trainlist) == len(elements):
            common = list(set.intersection(set(trainlist), set(elements)))
            diff1 = list(set.difference(set(trainlist), common)) #has to be deleted
            diff2 = list(set.difference(set(elements), common)) #to add
            # del
            for item in diff1:
                del analysis_results['analysis'][item]
            #add
            for item in diff2:
                analysis_results['analysis'][item] = {} #add

def read_location_entries():
    tracknumber = entry_track.get_text()
    track_loc = entry_track_loc.get_text()
    street = entry_street.get_text()
    postal = entry_postal.get_text()
    location = entry_location.get_text()
    state = combobox_state.get_active_text()
    country = combobox_country.get_active_text()
    return tracknumber, track_loc, street, postal, location, state, country

def location4textview(tracknumber, track_loc, street, postal, location, state, country):
    location_preview = []
    location_preview.append('Strecke:\t' + tracknumber +'\t' + 'Strecken-km: ' + track_loc)
    location_preview.append(street)
    location_preview.append(postal +'\t' + location)
    location_preview.append(state)
    location_preview.append(country)
    return location_preview

def location2textview(location_preview):
    textview_location_preview.set_property('editable', False)
    textview_location_preview.set_justification(Gtk.Justification.LEFT)
    buffer = textview_location_preview.get_buffer()
    buffer.delete(buffer.get_start_iter(), buffer.get_end_iter())
    string = ''
    for i in range(0, len(location_preview)):
        string_temp = '%s' % location_preview[i]
        string = string + string_temp + '\n'
    buffer.insert(buffer.get_end_iter(), string)

def get_notes():
    glob_notes = entry_notes.get_text()
    if glob_notes == '':
        glob_notes = 'keine'
    return glob_notes

def check_for_empty_entries(list):
    for item in list:
        if item == '':
            crit = True
            break
        else:
            crit = False
    return crit

def check_address_complete(tracknumber, track_loc, street, postal, location, state, country):
    if tracknumber != None and track_loc != None and street != None and postal != None and location != None and state != None and country != None:
        location_dialog.hide()
    else:
        warning_window('Fehler!: \nAdressangaben vervollständigen')

def warning_window(string):
    warnwindow = Gtk.Window()
    warnwindow.set_default_size(300, 100)
    hbox = Gtk.Box(spacing=2)
    warnwindow.add(hbox)  # label1 = Gtk.Label(str)
    label1 = Gtk.Label(string,
                       xalign=0.5)
    label1.set_justify(Gtk.Justification.CENTER)
    hbox.pack_start(label1, True, True, 0)
    warnwindow.show_all()

def check_configuration_complete(elements):
    if check_for_empty_entries(elements) == False and combobox_air_category.get_active_text() != None and combobox_landuse.get_active_text() != None and same_entries_in_list(elements) == False:
        configure_prognose.hide()
    elif check_for_empty_entries(elements) == False and combobox_air_category.get_active_text() != None and combobox_landuse.get_active_text() != None and same_entries_in_list(elements) == True:
        warning_window('Fehler!: \nDoppelbenennung eines Datensatzes')
    elif check_for_empty_entries(elements) == True and combobox_air_category.get_active_text() != None and combobox_landuse.get_active_text() != None:
        if same_entries_in_list(elements) == True:
            warning_window('Fehler!: \nBezeichnung von mindestens einem Datensatz leer \nDoppelbenennung eines Datensatzes')
        else:
            warning_window('Fehler!: \nBezeichnung von mindestens einem Datensatz leer')
    elif (combobox_air_category.get_active_text() == None or combobox_landuse.get_active_text() == None) and  check_for_empty_entries(elements) == False:
        warning_window('Fehler!: \nKeine Flächennutzung angegeben oder unvollständig')

    else:
        warning_window('Fehler!: \nBezeichnung von mindestens einem Datensatz leer. \nKeine Flächennutzung angegeben oder unvollständig')
    #print check_for_empty_entries(elements), combobox_air_category.get_active_text(), combobox_landuse.get_active_text(), same_entries_in_list(elements)

def same_entries_in_list(elements):
    if len(elements) == len(set(elements)):
        return False

    else:
        return True

def variables2results():
    analysis_results['analysis'][label] = {'MP5': MP5,'MP4': MP4,'MP3':MP3, 'MP2':MP2, 'MP1': MP1,'MP5Leq': MP5Leq,'MP4Leq': MP4Leq,'MP3Leq':MP3Leq, 'MP2Leq':MP2Leq, 'MP1Leq': MP1Leq}
    analysis_results['analysis'][label].update({'distMP5': distMP5, 'distMP4': distMP4, 'distBld': distBld,
                                             'track_shift': track_shift, 'LE': LE})
    analysis_results['analysis'][label].update({'vel_measure':vel_measure, 'vel_prognose_0':vel_prognose_0, 'vel_prognose_plan':vel_prognose_plan})
    if analysis_mode == 'zero':
        analysis_results['analysis'][label].update({'analysis_mode': analysis_mode, 'trains_day_0':trains_day_0, 'trains_night_0': trains_night_0, 'trains_day_plan': 0.0,
                                                  'trains_night_plan': 0.0, 'trainlength_0': trainlength_0, 'trainlength_plan': 0.0, 'v_inside_0': v_inside_0,
                                                  'v_inside_plan': np.zeros(20), 'Lv_inside_0': Lv_inside_0, 'Lv_inside_plan': np.zeros(20),'KBFtm_0': KBFtm_0, 'KBFtm_plan': 0.0,
                                                  'KBFtr_day_0': KBFtr_day_0, 'KBFtr_day_plan': 0.0, 'KBFtr_night_0': KBFtr_night_0,
                                                  'KBFtr_night_plan': 0.0, 'KBF_max_0': KBF_max_0, 'KBF_max_plan': 0.0,
                                                  'check4150_string_0': check4150_string_0, 'check4150_string_plan': '-',
                                                  'air_check_0': air_check_0, 'air_check_plan': '-', 'sig_KBF_max': 0.0,
                                                  'sig_KBFtm': 0.0, 'sig_KBFtr_day': 0.0, 'sig_KBFtr_night' :0.0,
                                                  'sig_Lm_day': 0.0, 'sig_Lm_night': 0.0, 'Lm_day_0': Lm_day_0, 'Lm_night_0': Lm_night_0,
                                                  'Lm_day_plan': -100.0,'Lm_night_plan': -100.0,'all_frequencies_calculation': all_frequencies_calculation})
    elif analysis_mode == 'plan':
        analysis_results['analysis'][label].update(
            {'analysis_mode': analysis_mode, 'trains_day_0': 0.0, 'trains_night_0': 0.0, 'trains_day_plan': trains_day_plan,
             'trains_night_plan': trains_night_plan, 'trainlength_0': 0.0, 'trainlength_plan': trainlength_plan, 'v_inside_0': np.zeros(20),
             'v_inside_plan': v_inside_plan, 'Lv_inside_0':np.zeros(20), 'Lv_inside_plan': Lv_inside_plan,
             'KBFtm_0': 0.0, 'KBFtm_plan': KBFtm_plan, 'KBFtr_day_0': 0.0, 'KBFtr_day_plan': KBFtr_day_plan,
             'KBFtr_night_0': 0.0, 'KBFtr_night_plan': KBFtr_night_plan, 'KBF_max_0': 0.0, 'KBF_max_plan': KBF_max_plan,
             'check4150_string_0': '-', 'check4150_string_plan': check4150_string_plan, 'air_check_0': '-',
             'air_check_plan': air_check_plan, 'sig_KBF_max': 0.0, 'sig_KBFtm': 0.0, 'sig_KBFtr_day': 0.0, 'sig_KBFtr_night': 0.0,
             'sig_Lm_day': 0.0, 'sig_Lm_night': 0.0, 'Lm_day_0': -100.0, 'Lm_night_0': -100.0,
             'Lm_day_plan': Lm_day_plan, 'Lm_night_plan': Lm_night_plan,
             'all_frequencies_calculation': all_frequencies_calculation})
    else:
        analysis_results['analysis'][label].update(
            {'analysis_mode': analysis_mode, 'trains_day_0': trains_day_0, 'trains_night_0': trains_night_0,
             'trains_day_plan': trains_day_plan,
             'trains_night_plan': trains_night_plan, 'trainlength_0': trainlength_0, 'trainlength_plan': trainlength_plan, 'v_inside_0': v_inside_0,
             'v_inside_plan': v_inside_plan, 'Lv_inside_0': Lv_inside_0, 'Lv_inside_plan': Lv_inside_plan,
             'KBFtm_0': KBFtm_0, 'KBFtm_plan': KBFtm_plan,
             'KBFtr_day_0': KBFtr_day_0, 'KBFtr_day_plan': KBFtr_day_plan, 'KBFtr_night_0': KBFtr_night_0,
             'KBFtr_night_plan': KBFtr_night_plan, 'KBF_max_0': KBF_max_0, 'KBF_max_plan': KBF_max_plan,
             'check4150_string_0': check4150_string_0, 'check4150_string_plan': check4150_string_plan,
             'air_check_0': air_check_0, 'air_check_plan': air_check_plan, 'sig_KBF_max': sig_KBF_max,
             'sig_KBFtm': sig_KBFtm, 'sig_KBFtr_day': sig_KBFtr_day, 'sig_KBFtr_night': sig_KBFtr_night,
             'sig_Lm_day': sig_Lm_day, 'sig_Lm_night': sig_Lm_night, 'Lm_day_0': Lm_day_0, 'Lm_night_0': Lm_night_0,
             'Lm_day_plan': Lm_day_plan, 'Lm_night_plan': Lm_night_plan,
             'all_frequencies_calculation': all_frequencies_calculation})
    #'trains_day':trains_day, 'trains_night':trains_night, 'trainlength': trainlength,
    #print analysis_results['analysis'][dataset].keys()
    #print Lv_outsideBld

def fill_buttons():
    traintypes = sorted(analysis_results['analysis'].keys())
    set_result_buttons_invisible()
    for i in range(0, len(traintypes)):
        button_traintype[i].set_label(traintypes[i])
        button_traintype[i].show_all()
        #button_del_traintype[i].show_all()

def set_result_buttons_invisible():
    for item in button_traintype:
        item.set_visible(False)
        item.hide()
    #for item in button_del_traintype:
        #item.set_visible(False)
        #item.hide()

def set_address_in_results():
    builder.get_object('label67').set_text('Strecke: ' + str(analysis_results['metadata']['tracknumber']) + '\t' + 'km: ' + str(analysis_results['metadata']['track_loc']))
    builder.get_object('label68').set_text(str(analysis_results['metadata']['street']))
    builder.get_object('label69').set_text(str(analysis_results['metadata']['postal'])+ ' ' + str(analysis_results['metadata']['location']))
    builder.get_object('label70').set_text(str(analysis_results['metadata']['state']))
    builder.get_object('label71').set_text(str(analysis_results['metadata']['country']))
    builder.get_object('label74').set_text(str(analysis_results['metadata']['landuse']))
    print_var2textview(analysis_results['metadata']['glob_notes'], 'textview38', 's')
    builder.get_object('textview38').set_justification(Gtk.Justification.LEFT)

def clear_all():
    # plots
    ax1.clear()
    common_ax1(ax1)
    sw1.queue_draw()
    ax2.clear()
    common_ax2(ax2)
    sw2.queue_draw()
    ax3.clear()
    common_ax3(ax3)
    sw3.queue_draw()
    ax4.clear()
    common_ax4(ax4)
    sw4.queue_draw()
    ax5.clear()
    common_ax5(ax5)
    sw5.queue_draw()
    ax6.clear()
    common_ax6(ax6)
    sw6.queue_draw()

    # textviews
    for i in range(1,24):
        textview = 'textview' + str(i)
        clear_textview(builder.get_object(textview))
    # entries
    builder.get_object('entry_MP5').set_text('')
    builder.get_object('entry_MP4').set_text('')
    builder.get_object('entry_dist_bld').set_text('')
    builder.get_object('entry_track_shift').set_text('')
    builder.get_object('entry_vel_measure').set_text('')
    builder.get_object('entry_vel_zero').set_text('')
    builder.get_object('entry_vel_plan').set_text('')
    builder.get_object('entry_trains_day_zero').set_text('')
    builder.get_object('entry_trains_night_zero').set_text('')
    builder.get_object('entry_trains_day_plan').set_text('')
    builder.get_object('entry_trains_night_plan').set_text('')
    builder.get_object('entry_trainlength_zero').set_text('')
    builder.get_object('entry_trainlength_plan').set_text('')
    # comboboxes

    #builder.get_object('comboboxtext2').set_active(-1)
    #builder.get_object('comboboxtext1').set_active(-1)

    #builder.get_object('combobox_dataset4analysis').set_active(-1)

    set_entries_passive()
    set_buttons_passive()
    set_comboboxes_passive()

    #clear all frequencies window
    clear_all_frequencies_window()

def clear_all_frequencies_window():
    for i in range(0, 22):
        for j in range(0, 22):
            text = all_frequencies.get_object(('text_' + str(i) + '_' + str(j)))
            text_opt_frqwin(text)
            buffer1 = text.get_buffer()
            buffer1.delete(buffer1.get_start_iter(), buffer1.get_end_iter())

def clear_textview(textview):
    buffer1 = textview.get_buffer()
    buffer1.delete(buffer1.get_start_iter(), buffer1.get_end_iter())

def print_var2textview(var, textviewname, precision):
    textview = builder.get_object(textviewname)
    text_opt(textview)
    buffer1 = textview.get_buffer()
    buffer1.delete(buffer1.get_start_iter(), buffer1.get_end_iter())
    string1 = ('%' + precision) % var
    buffer1.insert(buffer1.get_end_iter(), string1)

def print_string2textview(str, textviewname):
    textview = builder.get_object(textviewname)
    text_opt(textview)
    buffer1 = textview.get_buffer()
    buffer1.delete(buffer1.get_start_iter(), buffer1.get_end_iter())
    string1 = (str)
    buffer1.insert(buffer1.get_end_iter(), string1)

def check_for_analyzed_dataset():
    traintypes = analysis_results['analysis'].keys()
    traintypes_analyzed = []
    for item in traintypes:
        if analysis_results['analysis'][item].keys() != []:
            traintypes_analyzed.append(item)
    if traintypes_analyzed != []:
        return True
    else:
        return False

def calc_and_print_overall_results():
    traintypes_analyzed = []
    KBFmax_0_overall = []
    KBFmax_plan_overall = []
    KBFtr_day_0_overall = []
    KBFtr_day_plan_overall = []
    KBFtr_night_0_overall = []
    KBFtr_night_plan_overall = []
    Lm_day_0_overall = []
    Lm_day_plan_overall = []
    Lm_night_0_overall = []
    Lm_night_plan_overall = []

    traintypes = analysis_results['analysis'].keys()
    for item in traintypes:
        if analysis_results['analysis'][item].keys() != []:
            traintypes_analyzed.append(item)
    for item in traintypes_analyzed:
        KBFmax_0_overall.append(analysis_results['analysis'][item]['KBF_max_0'])
        KBFmax_plan_overall.append(analysis_results['analysis'][item]['KBF_max_plan'])
        KBFtr_day_0_overall.append(analysis_results['analysis'][item]['KBFtr_day_0'])
        KBFtr_day_plan_overall.append(analysis_results['analysis'][item]['KBFtr_day_plan'])
        KBFtr_night_0_overall.append(analysis_results['analysis'][item]['KBFtr_night_0'])
        KBFtr_night_plan_overall.append(analysis_results['analysis'][item]['KBFtr_night_plan'])
        Lm_day_0_overall.append(analysis_results['analysis'][item]['Lm_day_0'])
        Lm_day_plan_overall.append(analysis_results['analysis'][item]['Lm_day_plan'])
        Lm_night_0_overall.append(analysis_results['analysis'][item]['Lm_night_0'])
        Lm_night_plan_overall.append(analysis_results['analysis'][item]['Lm_night_plan'])

    KBFmax_0_overall = max(KBFmax_0_overall)
    KBFmax_plan_overall = max(KBFmax_plan_overall)
    delta_KBFmax = str('%4.2f' %(calc_percentage(KBFmax_0_overall, KBFmax_plan_overall))) + ' %'

    #print KBFtr_day_0_overall
    #print KBFtr_day_plan_overall

    KBFtr_day_0_overall = calc_sqrt_sqsum(KBFtr_day_0_overall)
    KBFtr_day_plan_overall = calc_sqrt_sqsum(KBFtr_day_plan_overall)
    KBFtr_night_0_overall = calc_sqrt_sqsum(KBFtr_night_0_overall)
    KBFtr_night_plan_overall = calc_sqrt_sqsum(KBFtr_night_plan_overall)
    delta_KBFtr_day = str('%4.2f' %(calc_percentage(KBFtr_day_0_overall, KBFtr_day_plan_overall))) +' %'
    delta_KBFtr_night = str('%4.2f' %(calc_percentage(KBFtr_night_0_overall, KBFtr_night_plan_overall))) + ' %'

    Lm_day_0_overall = energetic_level_addition(Lm_day_0_overall)
    Lm_day_plan_overall = energetic_level_addition(Lm_day_plan_overall)
    Lm_night_0_overall = energetic_level_addition(Lm_night_0_overall)
    Lm_night_plan_overall = energetic_level_addition(Lm_night_plan_overall)
    delta_Lm_day = str('%4.2f' %(Lm_day_plan_overall-Lm_day_0_overall)) +' db'
    delta_Lm_night = str('%4.2f' %(Lm_night_plan_overall - Lm_night_0_overall)) +' db'

    print_var2textview(KBFmax_0_overall, 'textview24', '6.4f')
    print_var2textview(KBFmax_plan_overall, 'textview25', '6.4f')
    print_var2textview(delta_KBFmax, 'textview39', 's')

    print_var2textview(KBFtr_day_0_overall, 'textview26', '6.4f')
    print_var2textview(KBFtr_day_plan_overall, 'textview29', '6.4f')
    print_var2textview(KBFtr_night_0_overall, 'textview27', '6.4f')
    print_var2textview(KBFtr_night_plan_overall, 'textview30', '6.4f')
    print_var2textview(delta_KBFtr_day, 'textview40', 's')
    print_var2textview(delta_KBFtr_night, 'textview41', 's')

    print_var2textview(Lm_day_0_overall, 'textview32', '6.4f')
    print_var2textview(Lm_day_plan_overall, 'textview35', '6.4f')
    print_var2textview(Lm_night_0_overall, 'textview33', '6.4f')
    print_var2textview(Lm_night_plan_overall, 'textview36', '6.4f')

    print_var2textview(delta_Lm_day, 'textview43', 's')
    print_var2textview(delta_Lm_night, 'textview44', 's')

    check4150_string_0_overall, check4150_string_plan_overall = check_4150_conditions(KBFmax_0_overall, KBFmax_plan_overall, KBFtr_day_0_overall,
                                                                                      KBFtr_night_0_overall, KBFtr_day_plan_overall, KBFtr_night_plan_overall,
                                                                                      analysis_results['metadata']['Au_day'], analysis_results['metadata']['Au_night'],
                                                                                      analysis_results['metadata']['Ao_day'], analysis_results['metadata']['Ao_night'],
                                                                                      analysis_results['metadata']['Ar_day'], analysis_results['metadata']['Ar_night'])

    print_var2textview(check4150_string_0_overall, 'textview28', 's')
    print_var2textview(check4150_string_plan_overall, 'textview31', 's')

    check_air_string_0_overall,check_air_string_plan_overall = check_sec_airborne_sound(analysis_results['metadata']['air_category'], Lm_day_0_overall, Lm_night_0_overall, Lm_day_plan_overall, Lm_night_plan_overall)

    print_var2textview(check_air_string_0_overall, 'textview34', 's')
    print_var2textview(check_air_string_plan_overall, 'textview37', 's')

    #print builder.get_object('switch1').get_state()

def plot_all_results():
    ax7.clear()
    traintypes = analysis_results['analysis'].keys()
    traintypes_analyzed = []
    for item in traintypes:
        if analysis_results['analysis'][item].keys() != []:
            traintypes_analyzed.append(item)
    color_idx = np.linspace(0, 2, 20)
    k=0
    for item in traintypes_analyzed:
        label = r'$\mathtt{%s}$' %item
        #label_1 = r'$\mathtt{L_{v-Raum-0}(f)-%s}$' % item
        if analysis_results['analysis'][item]['Lv_inside_0'].all() != 0:
            ax7.plot(ticks, analysis_results['analysis'][item]['Lv_inside_0'], label=label, color=cm.jet(color_idx[k]))
        if analysis_results['analysis'][item]['Lv_inside_plan'].all() != 0:
            ax7.plot(ticks, analysis_results['analysis'][item]['Lv_inside_plan'], color=cm.jet(color_idx[k]), linestyle='--')
        k =k+1
    ax7.legend(loc=0, fontsize=6)
    ax7.set_xticks(ticks, minor=False)
    ax7.set_xticklabels(labels_extended, rotation=45)
    ax7.xaxis.grid(True, which='major')
    ax7.set_title('Immissionsspektren im Gebäude', fontsize=9)
    ax7.set_xlabel('Frequenz [Hz]', fontsize=8)
    ax7.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize=8)
    ax7.tick_params(labelsize=7)
    ax7.set_xlim([0, 21])
    #ax7.set_ylim([-10, y_limits])

def energetic_level_addition(level_list):
    level = 10 * math.log10(np.sum(np.power(10, np.divide(level_list, 10))))
    return level

def calc_sqrt_sqsum(array):
    sol = math.sqrt(np.sum(np.power(array, 2)))
    return sol

def calc_percentage(p0,pplan):
    if p0 == 0.0 or pplan == 0.0:
        sol = np.nan
    else:
        sol = pplan*100/p0
    if sol >= 100:
        sol = sol - 100
    else:
        sol = -(100 -sol)
    return sol

def fill_results_single(label):
    builder.get_object('label90').set_text(label)
    if analysis_results['analysis'][label].keys() != []:
        if analysis_results['analysis'][label]['analysis_mode'] == 'zero':
            print_var2textview(analysis_results['analysis'][label]['KBF_max_0'], 'results_single_textview1', '6.4f')
            print_string2textview('-', 'results_single_textview2')
            print_string2textview('-', 'results_single_textview3')
            print_var2textview(analysis_results['analysis'][label]['KBFtm_0'], 'results_single_textview4', '6.4f')
            print_string2textview('-', 'results_single_textview5')
            print_string2textview('-', 'results_single_textview6')
            print_var2textview(analysis_results['analysis'][label]['KBFtr_day_0'], 'results_single_textview7', '6.4f')
            print_string2textview('-', 'results_single_textview8')
            print_string2textview('-', 'results_single_textview9')
            print_var2textview(analysis_results['analysis'][label]['KBFtr_night_0'], 'results_single_textview10', '6.4f')
            print_string2textview('-', 'results_single_textview11')
            print_string2textview('-', 'results_single_textview12')
            print_var2textview(analysis_results['analysis'][label]['check4150_string_0'], 'results_single_textview13', 's')
            print_string2textview('-', 'results_single_textview14')
            print_var2textview(analysis_results['analysis'][label]['Lm_day_0'], 'results_single_textview15', '6.4f')
            print_string2textview('-', 'results_single_textview16')
            print_string2textview('-', 'results_single_textview17')
            print_var2textview(analysis_results['analysis'][label]['Lm_night_0'], 'results_single_textview18', '6.4f')
            print_string2textview('-', 'results_single_textview19')
            print_string2textview('-', 'results_single_textview20')
            print_var2textview(analysis_results['analysis'][label]['air_check_0'], 'results_single_textview21', 's')
            print_string2textview('-', 'results_single_textview22')

        elif analysis_results['analysis'][label]['analysis_mode'] == 'plan':
            print_string2textview('-', 'results_single_textview1')
            print_var2textview(analysis_results['analysis'][label]['KBF_max_plan'], 'results_single_textview2', '6.4f')
            print_string2textview('-', 'results_single_textview3')
            print_string2textview('-', 'results_single_textview4')
            print_var2textview(analysis_results['analysis'][label]['KBFtm_plan'], 'results_single_textview5', '6.4f')
            print_string2textview('-', 'results_single_textview6')
            print_string2textview('-', 'results_single_textview7')
            print_var2textview(analysis_results['analysis'][label]['KBFtr_day_plan'], 'results_single_textview8','6.4f')
            print_string2textview('-', 'results_single_textview9')
            print_string2textview('-', 'results_single_textview10')
            print_var2textview(analysis_results['analysis'][label]['KBFtr_night_plan'], 'results_single_textview11','6.4f')
            print_string2textview('-', 'results_single_textview12')
            print_string2textview('-', 'results_single_textview13')
            print_var2textview(analysis_results['analysis'][label]['check4150_string_plan'],'results_single_textview14', 's')
            print_string2textview('-', 'results_single_textview15')
            print_var2textview(analysis_results['analysis'][label]['Lm_day_plan'], 'results_single_textview16', '6.4f')
            print_string2textview('-', 'results_single_textview17')
            print_string2textview('-', 'results_single_textview18')
            print_var2textview(analysis_results['analysis'][label]['Lm_night_plan'], 'results_single_textview19', '6.4f')
            print_string2textview('-', 'results_single_textview20')
            print_string2textview('-', 'results_single_textview21')
            print_var2textview(analysis_results['analysis'][label]['air_check_plan'], 'results_single_textview22', 's')
        else:
            print_var2textview(analysis_results['analysis'][label]['KBF_max_0'], 'results_single_textview1', '6.4f')
            print_var2textview(analysis_results['analysis'][label]['KBF_max_plan'], 'results_single_textview2', '6.4f')
            print_var2textview(calc_percentage(float(analysis_results['analysis'][label]['KBF_max_0']),
                                               float(analysis_results['analysis'][label]['KBF_max_plan'])),
                               'results_single_textview3', '4.2f')
            print_var2textview(analysis_results['analysis'][label]['KBFtm_0'], 'results_single_textview4', '6.4f')
            print_var2textview(analysis_results['analysis'][label]['KBFtm_plan'], 'results_single_textview5', '6.4f')
            print_var2textview(calc_percentage(float(analysis_results['analysis'][label]['KBFtm_0']),
                                               float(analysis_results['analysis'][label]['KBFtm_plan'])),
                               'results_single_textview6', '4.2f')
            print_var2textview(analysis_results['analysis'][label]['KBFtr_day_0'], 'results_single_textview7', '6.4f')
            print_var2textview(analysis_results['analysis'][label]['KBFtr_day_plan'], 'results_single_textview8',
                               '6.4f')
            print_var2textview(calc_percentage(float(analysis_results['analysis'][label]['KBFtr_day_0']),
                                               float(analysis_results['analysis'][label]['KBFtr_day_plan'])),
                               'results_single_textview9', '4.2f')
            print_var2textview(analysis_results['analysis'][label]['KBFtr_night_0'], 'results_single_textview10',
                               '6.4f')
            print_var2textview(analysis_results['analysis'][label]['KBFtr_night_plan'], 'results_single_textview11',
                               '6.4f')
            print_var2textview(calc_percentage(float(analysis_results['analysis'][label]['KBFtr_night_0']),
                                               float(analysis_results['analysis'][label]['KBFtr_night_plan'])),
                               'results_single_textview12', '4.2f')
            print_var2textview(analysis_results['analysis'][label]['check4150_string_0'], 'results_single_textview13',
                               's')
            print_var2textview(analysis_results['analysis'][label]['check4150_string_plan'],
                               'results_single_textview14', 's')
            print_var2textview(analysis_results['analysis'][label]['Lm_day_0'], 'results_single_textview15', '6.4f')
            print_var2textview(analysis_results['analysis'][label]['Lm_day_plan'], 'results_single_textview16', '6.4f')
            print_var2textview(str('%4.2f' % (float(analysis_results['analysis'][label]['Lm_day_plan']) - float(
                analysis_results['analysis'][label]['Lm_day_0']))) + ' db', 'results_single_textview17', 's')
            print_var2textview(analysis_results['analysis'][label]['Lm_night_0'], 'results_single_textview18', '6.4f')
            print_var2textview(analysis_results['analysis'][label]['Lm_night_plan'], 'results_single_textview19',
                               '6.4f')
            print_var2textview(str('%4.2f' % (float(analysis_results['analysis'][label]['Lm_night_plan']) - float(
                analysis_results['analysis'][label]['Lm_night_0']))) + ' db', 'results_single_textview20', 's')
            print_var2textview(analysis_results['analysis'][label]['air_check_0'], 'results_single_textview21', 's')
            print_var2textview(analysis_results['analysis'][label]['air_check_plan'], 'results_single_textview22', 's')

def clear_results_single():
    for i in range(1, 23):
        clear_textview(builder.get_object('results_single_textview' + str(i)))

def fill_prognose_data_single(label):
    builder.get_object('label92').set_text(label)
    print_var2textview(analysis_results['analysis'][label]['distMP5'], 'textview66', '4.2f')
    print_var2textview(analysis_results['analysis'][label]['distMP4'], 'textview67', '4.2f')
    print_var2textview(analysis_results['analysis'][label]['distBld'], 'textview68', '4.2f')
    print_var2textview(analysis_results['analysis'][label]['track_shift'], 'textview69', '4.2f')

    print_var2textview(analysis_results['analysis'][label]['vel_measure'], 'textview75', '4.2f')
    print_var2textview(analysis_results['analysis'][label]['vel_prognose_0'], 'textview76', '4.2f')
    print_var2textview(analysis_results['analysis'][label]['vel_prognose_plan'], 'textview77', '4.2f')

    print_var2textview(analysis_results['analysis'][label]['trains_day_0'], 'textview72', 'd')
    print_var2textview(analysis_results['analysis'][label]['trains_night_0'], 'textview70', 'd')
    print_var2textview(analysis_results['analysis'][label]['trains_day_plan'], 'textview73', 'd')
    print_var2textview(analysis_results['analysis'][label]['trains_night_plan'], 'textview71', 'd')
    # compability 4 version 1.05
    if 'trainlength_0' in analysis_results['analysis'][label].keys():
        print_var2textview(analysis_results['analysis'][label]['trainlength_0'], 'textview74', 'd')
        print_var2textview(analysis_results['analysis'][label]['trainlength_plan'], 'textview42', 'd')
    else:
        print_var2textview(analysis_results['analysis'][label]['trainlength'], 'textview74', 'd')
        print_var2textview(analysis_results['analysis'][label]['trainlength'], 'textview42', 'd')

def fill_configuration_from_file(builder):
    # location_dialog
    builder.get_object('entry_track').set_text(analysis_results['metadata']['tracknumber'])
    builder.get_object('entry_track_loc').set_text(analysis_results['metadata']['track_loc'])
    builder.get_object('entry_street').set_text(analysis_results['metadata']['street'])
    builder.get_object('entry_postal').set_text(analysis_results['metadata']['postal'])
    builder.get_object('entry_location').set_text(analysis_results['metadata']['location'])
    builder.get_object('combobox_state').set_active(set_item_in_comboboxtext(combobox_air_category, analysis_results['metadata']['state'],['Baden-Würtemberg', 'Bayern', 'Berlin', 'Brandenburg', 'Bremen', 'Hamburg', 'Hessen', 'Mecklenburg-Vorpommern', 'Niedersachsen', 'Nordrhein-Westfalen', 'Rheinland-Pfalz','Saarland', 'Sachsen', 'Sachsen-Anhalt', 'Schleswig-Holstein','Thüringen']))
    builder.get_object('combobox_country').set_active(set_item_in_comboboxtext(combobox_air_category, analysis_results['metadata']['country'], ['Deutschland']))
    Signals().location_apply(builder.get_object('button15'))


    # location dialog
    builder.get_object('entry_notes').set_text(analysis_results['metadata']['glob_notes'])
    builder.get_object('combobox_datasets').set_active(len(analysis_results['analysis'].keys())-1)
    for i in range(0, len(analysis_results['analysis'].keys())):
        dataset_entries[i].set_text(analysis_results['analysis'].keys()[i])
    builder.get_object('combobox_landuse').set_active(set_item_in_comboboxtext(combobox_landuse, analysis_results['metadata']['landuse'], ['Wohngebiet','Mischgebiet','Sondergebiet','Gewerbegebiet','Industriegebiet']))
    builder.get_object('combobox_category').set_active(set_item_in_comboboxtext(combobox_air_category, analysis_results['metadata']['air_category'],['Kat. 1', 'Kat. 2', 'Kat. 3', 'Kat. 4', 'Kat. 5','Kat. 6']))
    #print analysis_results['metadata']['OPNV']
    if 'OPNV' in analysis_results['metadata'].keys() == False:
        builder.get_object('switch1').set_state(False)
    else:
        builder.get_object('switch1').set_state(analysis_results['metadata']['OPNV'])
    Signals().apply_configuration(apply_button)

#######################################################################################################################
#######################################################################################################################
#####################################   DATABASE FUNCTIONS    #########################################################
#######################################################################################################################
#######################################################################################################################

def initialize_database_obj():
    database_obj = Gtk.Builder()
    database_obj.add_objects_from_file('./glade/db.glade', ('database_conversation', ''))
    database_obj.connect_signals(DB_Signals())
    return database_obj

def set_item_in_comboboxtext(combobox, item, itemlist):
    for i in range(0, len(itemlist)):
        if item == itemlist[i]:
            index = i
    return index

def initialize_database():
    connection = sqlite3.connect("V:/05_Fachthemen/02_Akustik/01_Erschütterung/Erschütterungsdatenbank/vibrations.db")
    # testing phase
    # connection = sqlite3.connect("vibrations.db")
    cursor = connection.cursor()
    return cursor

def where_arg4sql(argument):
    argument = "'" + argument + "'"
    return argument

def initialize_objects(database_obj):
    combobox_sorting = database_obj.get_object('combobox_sorting')
    label_1 = database_obj.get_object('label_1')
    combobox_1 = database_obj.get_object('combobox_1')
    label_2 = database_obj.get_object('label_2')
    combobox_2 = database_obj.get_object('combobox_2')
    label_3 = database_obj.get_object('label_3')
    combobox_3 = database_obj.get_object('combobox_3')
    label_4 = database_obj.get_object('label_4')
    combobox_4 = database_obj.get_object('combobox_4')
    label_5 = database_obj.get_object('label_5')
    combobox_5 = database_obj.get_object('combobox_5')
    return combobox_sorting, label_1, combobox_1, label_2, combobox_2, label_3, combobox_3, label_4, combobox_4, label_5, combobox_5

def fill_comboboxes_from_database(combobox, com):

    Gtk.CellLayout.clear(combobox)
    liststore = Gtk.ListStore(int, str)
    cursor.execute(com)
    result = cursor.fetchall()
    elements = []
    for r in result:
        elements.append(str(r[0]))
    elements = list(set(elements))
    for i in range(0, len(elements)):
        liststore.append([i, elements[i]])
    combobox.set_model(liststore)
    cell = Gtk.CellRendererText()
    combobox.pack_start(cell, True)
    combobox.add_attribute(cell, 'text', 1)

def fill_combobox_train_at_location(combobox, com):
    global db_ordinal
    Gtk.CellLayout.clear(combobox)
    liststore = Gtk.ListStore(int, str)
    cursor.execute(com)
    result = cursor.fetchall()
    elements = []
    db_ordinal = []
    for r in result:
        elements.append(str(r[0]) + ' ' + str(r[1]) + ' ' + str(r[2]) + ' @ ' + str(int(r[3])) + ' km/h')
        db_ordinal.append(r[4])
    for i in range(0, len(elements)):
        liststore.append([i, elements[i]])
    combobox.set_model(liststore)
    cell = Gtk.CellRendererText()
    combobox.pack_start(cell, True)
    combobox.add_attribute(cell, 'text', 1)

def initialize_scrolledwindow_DB(database_obj):
    sw_DB = database_obj.get_object('scrolledwindow_DB_conversation')
    fig_DB = Figure(figsize=(4,4), dpi=100)
    ax_DB = fig_DB.add_subplot(111)
    common_ax_DB(ax_DB)
    canvas = FigureCanvas(fig_DB)
    sw_DB.add_with_viewport(canvas)
    return sw_DB, fig_DB, ax_DB

def common_ax_DB(ax_DB):
    ax_DB.clear()
    ax_DB.set_xticks(ticks, minor=False)
    ax_DB.set_xticklabels(labels)
    ax_DB.xaxis.grid(True, which='major')
    ax_DB.set_title('Terzschnellespektren', fontsize=9)
    ax_DB.set_xlabel('Frequenz [Hz]', fontsize = 8)
    ax_DB.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize = 8)
    ax_DB.tick_params(labelsize =7)
    ax_DB.set_xlim([0, 21])

def clear_combobox(combobox):

    Gtk.CellLayout.clear(combobox)
    liststore = None
    combobox.set_model(liststore)
    cell = Gtk.CellRendererText()
    combobox.pack_start(cell, True)
    combobox.add_attribute(cell, 'text', 1)

def results2textview(statement):
    priority_window.get_object('textview1')
    com = "SELECT track_number,  FROM measurements WHERE " + statement

def get_measurements_from_DB(ordinal):
    com = "SELECT MP6, MP5, MP3, MP2, MP1, MP6_Leq, MP5_Leq, MP3_Leq, MP2_Leq, MP1_Leq, dist_MP6, dist_MP5 FROM measurements WHERE ordinal_number=%s" % (where_arg4sql(str(ordinal)))
    cursor.execute(com)
    result = cursor.fetchall()
    MP5 = sql_string2array(result[0][0])
    MP4 = sql_string2array(result[0][1])
    MP3 = sql_string2array(result[0][2])
    MP2 = sql_string2array(result[0][3])
    MP1 = sql_string2array(result[0][4])
    MP5Leq = sql_string2array(result[0][5])
    MP4Leq = sql_string2array(result[0][6])
    MP3Leq = sql_string2array(result[0][7])
    MP2Leq = sql_string2array(result[0][8])
    MP1Leq = sql_string2array(result[0][9])
    builder.get_object('entry_MP5').set_text(str(result[0][10]))
    builder.get_object('entry_MP4').set_text(str(result[0][11]))
    return MP5, MP4, MP3, MP2, MP1, MP5Leq, MP4Leq, MP3Leq, MP2Leq, MP1Leq

def sql_string2array(sql_string):
    res = sql_string.split()
    res = [float(i) for i in res]
    res = np.array(res)
    return res

def plot_measurement_from_database(ax_DB, sw_DB, ticks, MP5, MP4, MP3, MP2, MP1, MP5Leq, MP4Leq, MP3Leq, MP2Leq, MP1Leq, labels):
    ax_DB.clear()
    ax_DB.plot(ticks, MP5, label=r'$\mathtt{Emmission MaxHold}$', color='g', linewidth=1.5)
    ax_DB.plot(ticks, MP4, color='g', linewidth=1.5)
    ax_DB.plot(ticks, MP3, label=r'$\mathtt{Immission MaxHold}$', color='r', linewidth=1, linestyle='--')
    ax_DB.plot(ticks, MP2, color='r', linewidth=1, linestyle='--')
    ax_DB.plot(ticks, MP1, color='r', linewidth=1, linestyle='--')
    ax_DB.plot(ticks, MP5Leq, label=r'$\mathtt{Emmission Leq}$', color='m', linewidth=1.5)
    ax_DB.plot(ticks, MP4Leq, color='m', linewidth=1.5)
    ax_DB.plot(ticks, MP3Leq, label=r'$\mathtt{Immission Leq}$', color='c', linewidth=1, linestyle='--')
    ax_DB.plot(ticks, MP2Leq, color='c', linewidth=1, linestyle='--')
    ax_DB.plot(ticks, MP1Leq, color='c', linewidth=1, linestyle='--')
    ax_DB.legend(loc=0, fontsize=8)
    ax_DB.set_xticks(ticks, minor=False)
    ax_DB.set_xticklabels(labels)
    ax_DB.xaxis.grid(True, which='major')
    ax_DB.set_title('Emmissionsspektren der Messung', fontsize=9)
    ax_DB.set_xlabel('Frequenz [Hz]', fontsize=8)
    ax_DB.set_ylabel('Schwinggeschwindigkeit [dB]', fontsize=8)
    ax_DB.tick_params(labelsize=7)
    ax_DB.set_xlim([0, 21])
    ax_DB.set_ylim([0, 90])
    sw_DB.queue_draw()

def ylim_from_db_record(MP5, MP4, MP3, MP2, MP1, MP5Leq, MP4Leq, MP3Leq, MP2Leq, MP1Leq):
    maxi = []
    maxi.append(max(MP5))
    maxi.append(max(MP4))
    maxi.append(max(MP3))
    maxi.append(max(MP2))
    maxi.append(max(MP1))
    maxi.append(max(MP5Leq))
    maxi.append(max(MP4Leq))
    maxi.append(max(MP3Leq))
    maxi.append(max(MP2Leq))
    maxi.append(max(MP1Leq))
    y_limits = max(maxi)+10
    mod = y_limits % 10
    frac = 10 - mod
    y_limits = y_limits + frac
    return y_limits


#######################################################################################################################
#######################################################################################################################
terzfrequencies = [4, 5, 6.3, 8, 10, 12.5, 16, 20, 25, 31.5, 40, 50, 63, 80, 100, 125, 160, 200, 250, 315]
labels = ['', '5', '', '8', '', '12.5', '', '20', '', '31.5', '', '50', '', '80', '', '125', '', '200', '', '315']
labels_extended = ['4', '5', '6.3', '8', '10', '12.5', '16', '20', '25', '31.5', '40', '50', '63', '80', '100', '125', '160', '200', '250', '315']
ticks =np.arange(1,21,1.0)
model = False
transferfunction = None
analysis_results = {'metadata':{},'analysis':{}} #initialize

builder = initialize_builder()
analysis_window = builder.get_object('analysis_window')
configure_prognose = builder.get_object('configure_window')
location_dialog = builder.get_object('location_dialog')
main_window = builder.get_object('main_window')

dataset_entries = []

# initialize diagrams
############################################################
sw1, fig1, ax1 = initialize_scrolledwindow1(builder)
sw2, fig2, ax2 = initialize_scrolledwindow2(builder)
sw3, fig3, ax3 = initialize_scrolledwindow3(builder)
sw4, fig4, ax4 = initialize_scrolledwindow4(builder)
sw5, fig5, ax5 = initialize_scrolledwindow5(builder)
sw6, fig6, ax6 = initialize_scrolledwindow6(builder)
sw7, fig7, ax7 = initialize_scrolledwindow7(builder)
sw8, fig8, ax8 = initialize_scrolledwindow8(builder)
# initialize objects
############################################################
# main window
button_apply1, button_apply2, button_linear, button_Hoelzl, button_linear_model, button_Transferfunction_from_file, button_custom_tfbld, \
textview17, button_switch, button_modified_transferfunction, button_apply3, combobox_frequency, combobox_ceiling, menu_file_open,\
menu_file_from_database, combobox_dataset4analysis, button_write_variables2results\
    = initialize_main_window_objects(builder)
# location dialog
entry_track, entry_track_loc, entry_street, entry_postal, entry_location, combobox_state, combobox_country\
    = initialize_location_dialog_objects(builder)
# configure window
textview_location_preview, entry_notes, combobox_landuse, combobox_air_category \
    = initialize_configure_window_objects(builder)
# results window
button_traintype = initialize_result_window_buttons(builder)





# set entries and buttons passive
############################################################
set_entries_passive()
set_buttons_passive()
set_comboboxes_passive()
set_menuitems_passive()
set_result_buttons_invisible()
combobox_datasets = builder.get_object('combobox_datasets')

main_window.show_all()
set_result_buttons_invisible()
###################################################
#        not jet implemented functions            #
###################################################

button_DB_Richtlinie = builder.get_object('button4')
button_DB_Richtlinie.set_sensitive(False)

button_inversion = builder.get_object('button5')
button_inversion.set_sensitive(False)

button_mind1 = builder.get_object('button9')
button_mind1.set_sensitive(False)

button_mind2 = builder.get_object('button10')
button_mind2.set_sensitive(False)

button_mind3 = builder.get_object('button11')
button_mind3.set_sensitive(False)

button_BS = builder.get_object('button9')
button_BS.set_sensitive(False)
button_MFS = builder.get_object('button10')
button_MFS.set_sensitive(False)
button_BT = builder.get_object('button11')
button_BT.set_sensitive(False)

custom_reductions = initialize_custom_reductions_window()
custom_window = custom_reductions.get_object('custom_window')
custom_function = np.zeros(20)
reduction_type = 'None'

switch_dialog =builder.get_object('switch_dialog')

############ initialize all_frequencies_window
all_frequencies = initialize_view_all_frequencies_window()
all_frequencies_window = all_frequencies.get_object('all_frequencies_window')


###########################################################
###################### DATABASE ###########################
###########################################################

database_obj = initialize_database_obj()
cursor =initialize_database()
database_conversation_window = database_obj.get_object('database_conversation')
sw_DB, fig_DB, ax_DB = initialize_scrolledwindow_DB(database_obj)
combobox_sorting, label_1, combobox_1, label_2, combobox_2, label_3, combobox_3, label_4, combobox_4, label_5, combobox_5 = initialize_objects(database_obj)

def export2xls():
    wb = xl.load_workbook(workbook)
    new_name = str(len(wb.sheetnames))
    wb.create_sheet(new_name)
    instance = WorksheetCopy(wb.get_sheet_by_name('template'), wb.get_sheet_by_name(new_name))
    WorksheetCopy.copy_worksheet(instance)

    ws = wb[new_name]
    # metadata
    ws.cell(row=2, column=2).value = analysis_results['metadata']['tracknumber']
    ws.cell(row=2, column=5).value = analysis_results['metadata']['track_loc']
    ws.cell(row=2, column=9).value = label
    ws.cell(row=2, column=16).value = strftime("%Y-%m-%d %H:%M:%S", gmtime())
    ws.cell(row=3, column=2).value = analysis_results['metadata']['street']
    ws.cell(row=3, column=16).value = analysis_results['metadata']['glob_notes']
    ws.cell(row=4, column=2).value = analysis_results['metadata']['postal']
    ws.cell(row=4, column=5).value = analysis_results['metadata']['location']
    ws.cell(row=5, column=6).value = analysis_results['metadata']['landuse']
    #analysis data
    datalen = len(analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtm_0'])

    start_wood = 9
    for i in range (0, datalen/2):
        ws.cell(row=start_wood, column=2).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBF_max_0'][i]
        ws.cell(row=start_wood, column=3).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBF_max_plan'][i]
        ws.cell(row=start_wood, column=4).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_KBF_max'][i]
        ws.cell(row=start_wood, column=5).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtm_0'][i]
        ws.cell(row=start_wood, column=6).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtm_plan'][i]
        ws.cell(row=start_wood, column=7).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_KBFtm'][i]
        ws.cell(row=start_wood, column=8).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtr_day_0'][i]
        ws.cell(row=start_wood, column=9).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtr_day_plan'][i]
        ws.cell(row=start_wood, column=10).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_KBFtr_day'][i]
        ws.cell(row=start_wood, column=11).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtr_night_0'][i]
        ws.cell(row=start_wood, column=12).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtr_night_plan'][i]
        ws.cell(row=start_wood, column=13).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_KBFtr_night'][i]
        ws.cell(row=start_wood, column=14).value =analysis_results['analysis'][label]['all_frequencies_calculation']['all_Lm_day_0'][i]
        ws.cell(row=start_wood, column=15).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_Lm_day_plan'][i]
        ws.cell(row=start_wood, column=16).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_Lm_day'][i]
        ws.cell(row=start_wood, column=17).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_Lm_night_0'][i]
        ws.cell(row=start_wood, column=18).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_Lm_night_plan'][i]
        ws.cell(row=start_wood, column=19).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_Lm_night'][i]
        ws.cell(row=start_wood, column=20).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_check4150_string_0'][i]
        ws.cell(row=start_wood, column=21).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_check4150_string_plan'][i]
        ws.cell(row=start_wood, column=22).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_air_check_0'][i]
        ws.cell(row=start_wood, column=23).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_air_check_plan'][i]
        start_wood += 1
    start_concrete =24
    for i in range(datalen / 2, datalen):
        ws.cell(row=start_concrete, column=2).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBF_max_0'][i]
        ws.cell(row=start_concrete, column=3).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBF_max_plan'][i]
        ws.cell(row=start_concrete, column=4).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_KBF_max'][i]
        ws.cell(row=start_concrete, column=5).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtm_0'][i]
        ws.cell(row=start_concrete, column=6).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtm_plan'][i]
        ws.cell(row=start_concrete, column=7).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_KBFtm'][i]
        ws.cell(row=start_concrete, column=8).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtr_day_0'][i]
        ws.cell(row=start_concrete, column=9).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtr_day_plan'][i]
        ws.cell(row=start_concrete, column=10).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_KBFtr_day'][i]
        ws.cell(row=start_concrete, column=11).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtr_night_0'][i]
        ws.cell(row=start_concrete, column=12).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_KBFtr_night_plan'][i]
        ws.cell(row=start_concrete, column=13).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_KBFtr_night'][i]
        ws.cell(row=start_concrete, column=14).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_Lm_day_0'][i]
        ws.cell(row=start_concrete, column=15).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_Lm_day_plan'][i]
        ws.cell(row=start_concrete, column=16).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_Lm_day'][i]
        ws.cell(row=start_concrete, column=17).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_Lm_night_0'][i]
        ws.cell(row=start_concrete, column=18).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_Lm_night_plan'][i]
        ws.cell(row=start_concrete, column=19).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_sig_Lm_night'][i]
        ws.cell(row=start_concrete, column=20).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_check4150_string_0'][i]
        ws.cell(row=start_concrete, column=21).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_check4150_string_plan'][i]
        ws.cell(row=start_concrete, column=22).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_air_check_0'][i]
        ws.cell(row=start_concrete, column=23).value = analysis_results['analysis'][label]['all_frequencies_calculation']['all_air_check_plan'][i]
        start_concrete += 1

    wb.save(workbook)
    # ws.cell(row=2, column=2).font = Font(bold=True, vertAlign='subscript')


workbook = 'D:/work/Auswertung 17-53480/2030/results.xlsx'




Gtk.main()